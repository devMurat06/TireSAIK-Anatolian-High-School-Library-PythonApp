import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os
import csv
import shutil
from collections import Counter
from datetime import datetime, timedelta
import random
import string
import sys
import threading
import time
import json
import urllib.request

# macOS için zbar kütüphane yolu (pyzbar için gerekli)
if sys.platform == 'darwin':
    # Apple Silicon ve Intel Mac desteği
    zbar_paths = [
        '/opt/homebrew/lib',  # Apple Silicon
        '/usr/local/lib',      # Intel Mac
        '/opt/homebrew/Cellar/zbar/0.23.93_2/lib'  # Spesifik versiyon
    ]
    existing_path = os.environ.get('DYLD_LIBRARY_PATH', '')
    new_paths = ':'.join([p for p in zbar_paths if os.path.exists(p)])
    if new_paths:
        os.environ['DYLD_LIBRARY_PATH'] = new_paths + ':' + existing_path
        os.environ['DYLD_FALLBACK_LIBRARY_PATH'] = new_paths + ':' + existing_path

# Opsiyonel kütüphaneler
try:
    from openpyxl import load_workbook
    EXCEL_DESTEGI = True
except ImportError:
    EXCEL_DESTEGI = False

try:
    import barcode
    from barcode.writer import ImageWriter
    BARKOD_OLUSTURMA_DESTEGI = True
except ImportError:
    BARKOD_OLUSTURMA_DESTEGI = False

# OpenCV ve pyzbar için özel import
BARKOD_OKUMA_DESTEGI = False
try:
    import cv2
    try:
        from pyzbar import pyzbar
        BARKOD_OKUMA_DESTEGI = True
    except Exception as e:
        print(f"pyzbar yüklenemedi (zbar eksik olabilir): {e}")
except ImportError:
    pass

try:
    from PIL import Image, ImageTk
    PIL_DESTEGI = True
except ImportError:
    PIL_DESTEGI = False

try:
    import customtkinter as ctk
    CTK_DESTEGI = True
except ImportError:
    CTK_DESTEGI = False

# PDF desteği
try:
    from fpdf import FPDF
    PDF_DESTEGI = True
except ImportError:
    PDF_DESTEGI = False

# Matplotlib desteği (grafikler için)
try:
    import matplotlib
    matplotlib.use('TkAgg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_DESTEGI = True
except ImportError:
    MATPLOTLIB_DESTEGI = False

class GirisEkrani:
    """Öğretmen ve öğrenci giriş ekranı - Modern CTK"""
    
    def __init__(self, root, on_giris_basarili):
        self.root = root
        self.on_giris_basarili = on_giris_basarili
        self.parolalari_yukle()
        
        self.root.title("ŞAİK Kütüphane")
        self.root.geometry("400x450")
        self.root.resizable(False, False)
        
        if CTK_DESTEGI:
            self.root.configure(fg_color="#1a1a2e")
            self._build_ctk_ui()
        else:
            self.root.configure(bg="#1a1a2e")
            self._build_tk_ui()
    
    def _build_ctk_ui(self):
        """CustomTkinter modern UI"""
        main = ctk.CTkFrame(self.root, fg_color="#1a1a2e")
        main.pack(expand=True, fill="both", padx=40, pady=40)
        
        ctk.CTkLabel(main, text="📚", font=("Arial", 52), text_color="white").pack(pady=(20, 10))
        ctk.CTkLabel(main, text="ŞAİK KÜTÜPHANE", font=("Arial", 22, "bold"), text_color="#3b82f6").pack()
        ctk.CTkLabel(main, text="Yönetim Sistemi", font=("Arial", 12), text_color="#888").pack(pady=(0, 30))
        
        ctk.CTkButton(main, text="👨‍🏫  ÖĞRETMEN GİRİŞİ", font=("Arial", 14, "bold"),
                      width=250, height=50, corner_radius=10,
                      fg_color="#2563eb", hover_color="#1d4ed8",
                      command=lambda: self.parola_sor("ogretmen")).pack(pady=10)
        
        ctk.CTkButton(main, text="👨‍🎓  ÖĞRENCİ GİRİŞİ", font=("Arial", 14, "bold"),
                      width=250, height=50, corner_radius=10,
                      fg_color="#0891b2", hover_color="#0e7490",
                      command=lambda: self.parola_sor("ogrenci")).pack(pady=10)
        
        ctk.CTkLabel(main, text="© 2026 ŞAİK", font=("Arial", 10), text_color="#555").pack(side="bottom", pady=10)
    
    def _build_tk_ui(self):
        """Fallback Tkinter UI"""
        main = tk.Frame(self.root, bg="#1a1a2e")
        main.pack(expand=True, fill="both", padx=30, pady=30)
        tk.Label(main, text="📚", font=("Arial", 42), bg="#1a1a2e", fg="white").pack(pady=(10, 5))
        tk.Label(main, text="ŞAİK KÜTÜPHANE", font=("Arial", 20, "bold"), bg="#1a1a2e", fg="#3b82f6").pack()
        tk.Button(main, text="ÖĞRETMEN GİRİŞİ", font=("Arial", 12, "bold"), bg="#2563eb", fg="white",
                  width=22, height=2, command=lambda: self.parola_sor("ogretmen")).pack(pady=10)
        tk.Button(main, text="ÖĞRENCİ GİRİŞİ", font=("Arial", 12, "bold"), bg="#0891b2", fg="white",
                  width=22, height=2, command=lambda: self.parola_sor("ogrenci")).pack(pady=10)
    
    def parolalari_yukle(self):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base_dir, "okul_kutuphanesi_pro_v7.db")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("""CREATE TABLE IF NOT EXISTS parolalar (tip TEXT PRIMARY KEY, parola TEXT NOT NULL)""")
        cursor.execute("INSERT OR IGNORE INTO parolalar VALUES ('ogretmen', 'saik2026')")
        cursor.execute("INSERT OR IGNORE INTO parolalar VALUES ('ogrenci', 'ogrenci+')")
        conn.commit()
        cursor.execute("SELECT parola FROM parolalar WHERE tip='ogretmen'")
        self.OGRETMEN_PAROLA = cursor.fetchone()[0]
        cursor.execute("SELECT parola FROM parolalar WHERE tip='ogrenci'")
        self.OGRENCI_PAROLA = cursor.fetchone()[0]
        conn.close()
    
    def parola_sor(self, kullanici_tipi):
        if CTK_DESTEGI:
            dialog = ctk.CTkToplevel(self.root)
            dialog.title("Parola")
            dialog.geometry("320x200")
            dialog.resizable(False, False)
            dialog.transient(self.root)
            dialog.grab_set()
            
            baslik = "Öğretmen Parolası" if kullanici_tipi == "ogretmen" else "Öğrenci Parolası"
            ctk.CTkLabel(dialog, text=baslik, font=("Arial", 14, "bold")).pack(pady=25)
            
            parola_entry = ctk.CTkEntry(dialog, show="*", font=("Arial", 14), width=200, justify="center")
            parola_entry.pack(pady=10)
            parola_entry.focus_set()
            
            def giris_yap(event=None):
                if parola_entry.get() == (self.OGRETMEN_PAROLA if kullanici_tipi == "ogretmen" else self.OGRENCI_PAROLA):
                    dialog.destroy()
                    self.on_giris_basarili(kullanici_tipi)
                else:
                    messagebox.showerror("Hata", "Yanlış parola!", parent=dialog)
                    parola_entry.delete(0, "end")
            
            parola_entry.bind("<Return>", giris_yap)
            ctk.CTkButton(dialog, text="GİRİŞ", font=("Arial", 12, "bold"), width=120,
                          fg_color="#10b981", hover_color="#059669", command=giris_yap).pack(pady=15)
        else:
            # Fallback
            parola_pencere = tk.Toplevel(self.root)
            parola_pencere.title("Parola")
            parola_pencere.geometry("300x180")
            parola_pencere.configure(bg="#16213e")
            baslik = "Öğretmen Parolası" if kullanici_tipi == "ogretmen" else "Öğrenci Parolası"
            tk.Label(parola_pencere, text=baslik, font=("Arial", 12, "bold"), bg="#16213e", fg="white").pack(pady=20)
            parola_entry = tk.Entry(parola_pencere, show="*", font=("Arial", 12), width=20, justify="center")
            parola_entry.pack(pady=5)
            parola_entry.focus_set()
            def giris_yap(event=None):
                if parola_entry.get() == (self.OGRETMEN_PAROLA if kullanici_tipi == "ogretmen" else self.OGRENCI_PAROLA):
                    parola_pencere.destroy()
                    self.on_giris_basarili(kullanici_tipi)
                else:
                    messagebox.showerror("Hata", "Yanlış parola!", parent=parola_pencere)
            parola_entry.bind("<Return>", giris_yap)
            tk.Button(parola_pencere, text="GİRİŞ", bg="#10b981", fg="white", command=giris_yap).pack(pady=15)


class KutuphaneUygulamasi:
    def __init__(self, root, kullanici_tipi="ogretmen"):
        self.root = root
        self.kullanici_tipi = kullanici_tipi  # "ogretmen" veya "ogrenci"
        
        baslik = "ŞAİK Kütüphane Yönetim Sistemi"
        if kullanici_tipi == "ogrenci":
            baslik += " (Öğrenci Modu)"
        self.root.title(baslik)
        self.root.geometry("1280x768")

        # --- İKON AYARI ---
        self.uygulama_ikonu_ayarla()

        # --- MODERN RENK PALETİ ---
        self.bg_color = "#f8fafc"        # Açık gri-beyaz arka plan
        self.panel_color = "#1e293b"     # Koyu lacivert sol panel
        self.accent_color = "#3b82f6"    # Mavi vurgu
        self.action_color = "#10b981"    # Yeşil (başarı/işlem)
        self.danger_color = "#ef4444"    # Kırmızı (tehlike)
        self.warning_color = "#f59e0b"   # Turuncu (uyarı)
        self.text_primary = "#1e293b"    # Ana metin
        self.text_secondary = "#64748b"  # İkincil metin
        self.card_bg = "#ffffff"         # Kart arka planı
        self.border_color = "#e2e8f0"    # Kenarlık
        
        self.root.configure(bg=self.bg_color)

        # --- MODERN STİL AYARLARI ---
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Genel stiller
        self.style.configure("TFrame", background=self.bg_color)
        self.style.configure("TLabel", background=self.panel_color, foreground="white", font=("Segoe UI", 10))
        self.style.configure("TButton", font=("Segoe UI", 10, "bold"), borderwidth=0, focuscolor="none", padding=8)
        self.style.configure("TEntry", padding=8, font=("Segoe UI", 10))
        
        # Buton stilleri - Modern, yuvarlak görünüm
        self.style.configure("Ekle.TButton", background=self.accent_color, foreground="white", padding=10)
        self.style.map("Ekle.TButton", background=[('active', '#2563eb'), ('pressed', '#1d4ed8')])
        
        self.style.configure("Islem.TButton", background=self.action_color, foreground="white", font=("Segoe UI", 11, "bold"), padding=12)
        self.style.map("Islem.TButton", background=[('active', '#059669'), ('pressed', '#047857')])
        
        self.style.configure("Sil.TButton", background=self.danger_color, foreground="white", padding=10)
        self.style.map("Sil.TButton", background=[('active', '#dc2626'), ('pressed', '#b91c1c')])
        
        self.style.configure("Rapor.TButton", background=self.warning_color, foreground="white", padding=10)
        self.style.map("Rapor.TButton", background=[('active', '#d97706'), ('pressed', '#b45309')])
        
        self.style.configure("Normal.TButton", background="#64748b", foreground="white", padding=8)
        self.style.map("Normal.TButton", background=[('active', '#475569')])
        
        self.style.configure("Devre.TButton", background="#cbd5e1", foreground="#94a3b8", padding=8)
        
        # Treeview modern stili
        self.style.configure("Treeview", 
                            background=self.card_bg,
                            foreground=self.text_primary,
                            fieldbackground=self.card_bg,
                            rowheight=32,
                            font=("Segoe UI", 10))
        self.style.configure("Treeview.Heading", 
                            background="#f1f5f9",
                            foreground=self.text_primary,
                            font=("Segoe UI", 10, "bold"),
                            padding=8)
        self.style.map("Treeview", 
                       background=[('selected', '#dbeafe')],
                       foreground=[('selected', '#1e40af')])

        # Veritabanı Başlat
        self.db_adi = "okul_kutuphanesi_pro_v7.db"
        self.veritabani_kur()

        # Üst Menü
        self.menu_olustur()

        # Arayüzü Kur
        self.arayuz_olustur()
        self.verileri_guncelle()
        
        self.context_menu = tk.Menu(self.root, tearoff=0)

    # --- İKON AYARLAMA ---
    def uygulama_ikonu_ayarla(self):
        try:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            logo_path = os.path.join(base_dir, "logo.png")
            if os.path.exists(logo_path):
                img = tk.PhotoImage(file=logo_path)
                self.root.iconphoto(False, img)
                try:
                    import ctypes
                    myappid = 'saik.kutuphane.yonetim.v7'
                    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
                except: pass
        except Exception as e: print(f"İkon hatası: {e}")

    def tr_upper(self, text):
        if not text: return ""
        text = str(text)
        tr_map = {'ç': 'c~', 'Ç': 'C~', 'ğ': 'g~', 'Ğ': 'G~', 'ı': 'h~', 'I': 'H~',
                  'i': 'i', 'İ': 'I~~', 'ö': 'o~', 'Ö': 'O~', 'ş': 's~', 'Ş': 'S~', 'ü': 'u~', 'Ü': 'U~'}
        for key, val in tr_map.items(): text = text.replace(key, val)
        return text.lower()

    def veritabani_kur(self):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base_dir, self.db_adi)
        self.conn = sqlite3.connect(db_path)
        self.cursor = self.conn.cursor()
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS kitaplar (id INTEGER PRIMARY KEY AUTOINCREMENT, ad TEXT NOT NULL, yazar TEXT NOT NULL, tur TEXT, sayfa_sayisi INTEGER, raf_no TEXT, durum TEXT DEFAULT 'Mevcut', barkod TEXT UNIQUE, adet INTEGER DEFAULT 1)""")
        
        # Barkod ve adet sütunları yoksa ekle (eski veritabanları için)
        try:
            self.cursor.execute("ALTER TABLE kitaplar ADD COLUMN barkod TEXT UNIQUE")
            self.conn.commit()
        except sqlite3.OperationalError:
            pass
        try:
            self.cursor.execute("ALTER TABLE kitaplar ADD COLUMN adet INTEGER DEFAULT 1")
            self.conn.commit()
        except sqlite3.OperationalError:
            pass
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS odunc_alanlar (id INTEGER PRIMARY KEY AUTOINCREMENT, kitap_id INTEGER, ogrenci_ad TEXT, ogrenci_no TEXT, sinif TEXT, alinma_tarihi TEXT, iade_tarihi TEXT, FOREIGN KEY(kitap_id) REFERENCES kitaplar(id))""")
        
        # --- GEÇMİŞ TABLOSU (WRAPPED İÇİN) ---
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS odunc_gecmisi (id INTEGER PRIMARY KEY AUTOINCREMENT, kitap_ad TEXT, yazar TEXT, tur TEXT, ogrenci_ad TEXT, ogrenci_no TEXT, sinif TEXT, alinma_tarihi TEXT, iade_tarihi TEXT)""")
        
        # ogrenci_no sütunu yoksa ekle
        try:
            self.cursor.execute("ALTER TABLE odunc_gecmisi ADD COLUMN ogrenci_no TEXT")
            self.conn.commit()
        except sqlite3.OperationalError:
            pass
        
        # --- ÖĞRENCİLER TABLOSU ---
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS ogrenciler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            okul_no TEXT UNIQUE NOT NULL,
            ad_soyad TEXT NOT NULL,
            sinif TEXT,
            toplam_puan INTEGER DEFAULT 0,
            toplam_kitap INTEGER DEFAULT 0
        )""")
        
        # Puan sütunları yoksa ekle
        try:
            self.cursor.execute("ALTER TABLE ogrenciler ADD COLUMN toplam_puan INTEGER DEFAULT 0")
            self.cursor.execute("ALTER TABLE ogrenciler ADD COLUMN toplam_kitap INTEGER DEFAULT 0")
            self.conn.commit()
        except sqlite3.OperationalError:
            pass
        
        # --- REZERVASYON TABLOSU ---
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS rezervasyonlar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kitap_id INTEGER,
            ogrenci_no TEXT,
            ogrenci_ad TEXT,
            tarih TEXT,
            durum TEXT DEFAULT 'Bekliyor',
            FOREIGN KEY(kitap_id) REFERENCES kitaplar(id)
        )""")
        
        # --- ROZETLER TABLOSU ---
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS rozetler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ogrenci_no TEXT,
            rozet_adi TEXT,
            rozet_aciklama TEXT,
            kazanim_tarihi TEXT
        )""")
        
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS ayarlar (anahtar TEXT PRIMARY KEY, deger TEXT)""")
        self.cursor.execute("INSERT OR IGNORE INTO ayarlar (anahtar, deger) VALUES ('odunc_suresi', '45')")
        self.cursor.execute("INSERT OR IGNORE INTO ayarlar (anahtar, deger) VALUES ('tema', 'dark')")
        self.cursor.execute("INSERT OR IGNORE INTO ayarlar (anahtar, deger) VALUES ('otomatik_yedekleme', '1')")
        
        # --- YORUMLAR TABLOSU (Kitap önerileri için) ---
        self.cursor.execute("""CREATE TABLE IF NOT EXISTS yorumlar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kitap_id INTEGER,
            ogrenci_ad TEXT,
            ogrenci_no TEXT,
            yorum TEXT,
            puan INTEGER DEFAULT 5,
            tarih TEXT,
            FOREIGN KEY(kitap_id) REFERENCES kitaplar(id)
        )""")
        
        self.conn.commit()
        
        # Otomatik yedekleme başlat
        self.otomatik_yedekleme_baslat()

    def menu_olustur(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        dosya_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Dosya", menu=dosya_menu)
        
        # Excel Import (herkes için)
        if EXCEL_DESTEGI:
            dosya_menu.add_command(label="📥 Excel'den Kitap Aktar", command=self.excel_import)
        else:
            dosya_menu.add_command(label="📥 Excel'den Kitap Aktar (openpyxl yükleyin)", state="disabled")
        dosya_menu.add_separator()
        
        dosya_menu.add_command(label="💾 Veritabanını Yedekle", command=self.yedekle)
        
        # Yıllık Özet (sadece öğretmen)
        if self.kullanici_tipi == "ogretmen":
            dosya_menu.add_command(label="🏆 Yıllık Özet", command=self.wrapped_penceresi)
        
        dosya_menu.add_separator()
        
        # PDF Rapor
        if PDF_DESTEGI:
            dosya_menu.add_command(label="📄 PDF Rapor Oluştur", command=self.pdf_rapor_olustur)
        else:
            dosya_menu.add_command(label="📄 PDF Rapor (fpdf2 yükleyin)", state="disabled")
        
        dosya_menu.add_command(label="📊 CSV Rapor Al", command=self.rapor_al)
        dosya_menu.add_separator()
        dosya_menu.add_command(label="🚪 Çıkış Yap", command=self.cikis_yap)
        
        # İstatistikler Menüsü (sadece öğretmen)
        if self.kullanici_tipi == "ogretmen":
            istatistik_menu = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="📈 İstatistikler", menu=istatistik_menu)
            istatistik_menu.add_command(label="📊 İstatistik Dashboard", command=self.istatistik_dashboard)
            istatistik_menu.add_command(label="🏆 Sınıf Sıralaması", command=self.sinif_siralamasi)
            istatistik_menu.add_command(label="👑 Kitap Kurdu Liderlik Tablosu", command=self.liderlik_tablosu)
        
        # Ayarlar Menüsü (sadece öğretmen)
        if self.kullanici_tipi == "ogretmen":
            ayar_menu = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="⚙️ Ayarlar", menu=ayar_menu)
            ayar_menu.add_command(label="🔑 Parola Değiştir", command=self.parola_degistir_penceresi)
            ayar_menu.add_command(label="📅 Ödünç Süresi", command=self.ayarlar_penceresi)
            ayar_menu.add_separator()
            ayar_menu.add_command(label="🎨 Tema Değiştir", command=self.tema_degistir)
            ayar_menu.add_separator()
            ayar_menu.add_command(label="⏰ Gecikme Bildirimi Gönder", command=self.gecikme_bildirimi_goster)
            
        # Yapay Zeka Asistanı (sadece öğretmen)
        if self.kullanici_tipi == "ogretmen":
            ai_menu = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="🤖 AI Asistan", menu=ai_menu)
            ai_menu.add_command(label="💬 Asistan ile Sohbet", command=self.ai_asistan_penceresi)

        
        # Barkod Menüsü
        barkod_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="📊 Barkod İşlemleri", menu=barkod_menu)
        
        if self.kullanici_tipi == "ogretmen":
            if BARKOD_OLUSTURMA_DESTEGI:
                barkod_menu.add_command(label="🏷️ Seçili Kitap İçin Barkod Oluştur", command=self.barkod_olustur)
                barkod_menu.add_command(label="🏷️ Tüm Kitaplara Barkod Oluştur", command=self.toplu_barkod_olustur)
            else:
                barkod_menu.add_command(label="🏷️ Barkod Oluştur (python-barcode yükleyin)", state="disabled")
            barkod_menu.add_separator()
        
        if BARKOD_OKUMA_DESTEGI:
            barkod_menu.add_command(label="📷 Barkod Tara (Kamera)", command=self.barkod_tara)
        else:
            barkod_menu.add_command(label="📷 Barkod Tara (opencv-python & pyzbar yükleyin)", state="disabled")
        
        barkod_menu.add_separator()
        barkod_menu.add_command(label="🔍 Barkod ile Ara", command=self.barkod_ile_ara)
        
        # Öğrenciler Menüsü
        ogrenci_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="👥 Öğrenciler", menu=ogrenci_menu)
        
        if EXCEL_DESTEGI:
            ogrenci_menu.add_command(label="📥 Excel'den Öğrenci Aktar", command=self.ogrenci_excel_import)
        else:
            ogrenci_menu.add_command(label="📥 Excel'den Öğrenci Aktar (openpyxl yükleyin)", state="disabled")
        
        ogrenci_menu.add_separator()
        ogrenci_menu.add_command(label="👥 Öğrenci Listesi", command=self.ogrenci_listesi_penceresi)
        ogrenci_menu.add_command(label="🏅 Rozetler ve Puanlar", command=self.rozet_yonetimi)
        
        # Rezervasyon Menüsü
        rezervasyon_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="📅 Rezervasyonlar", menu=rezervasyon_menu)
        rezervasyon_menu.add_command(label="📋 Aktif Rezervasyonlar", command=self.rezervasyon_listesi)
        if self.kullanici_tipi == "ogretmen":
            rezervasyon_menu.add_command(label="➕ Yeni Rezervasyon", command=self.yeni_rezervasyon)
        
        # Oyunlar Menüsü (öğrenci modu için eğlence)
        oyun_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="🎮 Oyunlar", menu=oyun_menu)
        oyun_menu.add_command(label="🏓 Pong", command=self.oyun_pong)
        oyun_menu.add_command(label="🐦 Flappy Bird", command=self.oyun_flappy)
        oyun_menu.add_command(label="🐍 Yılan", command=self.oyun_yilan)
        
        # Alıştırmalar Menüsü (beyin egzersizleri)
        alistirma_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="🧠 Alıştırmalar", menu=alistirma_menu)
        alistirma_menu.add_command(label="🔢 Hızlı Matematik", command=self.alistirma_matematik)
        alistirma_menu.add_command(label="🔤 İngilizce Kelime", command=self.alistirma_kelime)
        alistirma_menu.add_command(label="🧩 Hafıza Oyunu", command=self.alistirma_hafiza)
        alistirma_menu.add_command(label="📝 Harf Karıştırma", command=self.alistirma_karistirma)
        alistirma_menu.add_command(label="⚡ Hızlı Tepki", command=self.alistirma_tepki)

    def arayuz_olustur(self):
        # SOL PANEL - Modern sidebar
        left_panel = tk.Frame(self.root, bg=self.panel_color, width=300)
        left_panel.pack(side=tk.LEFT, fill=tk.Y)
        left_panel.pack_propagate(False)

        # Logo / Başlık
        header_frame = tk.Frame(left_panel, bg=self.panel_color)
        header_frame.pack(pady=30)

        tk.Label(header_frame, text="📚", bg=self.panel_color, font=("Arial", 36)).pack()
        tk.Label(header_frame, text="ŞAİK", bg=self.panel_color, fg="white", font=("Segoe UI", 24, "bold")).pack()
        tk.Label(header_frame, text="KÜTÜPHANE", bg=self.panel_color, fg="#94a3b8", font=("Segoe UI", 11)).pack()

        # İstatistik Kartları - Modern görünüm
        stats_frame = tk.Frame(left_panel, bg="#334155", pady=15)
        stats_frame.pack(fill=tk.X, padx=15, pady=15)
        
        # Her stat için mini kart
        stat_items = [
            ("📚", "Toplam:", "white", "lbl_toplam"),
            ("📖", "Ödünçte:", "#fbbf24", "lbl_odunc"),
            ("⚠️", "Gecikmiş:", "#f87171", "lbl_gecikmis")
        ]
        
        self.lbl_toplam_kitap = tk.Label(stats_frame, text="📚 Toplam: 0", bg="#334155", fg="white", font=("Segoe UI", 10, "bold"))
        self.lbl_toplam_kitap.pack(anchor="w", padx=12, pady=3)
        self.lbl_odunc_kitap = tk.Label(stats_frame, text="📖 Ödünçte: 0", bg="#334155", fg="#fbbf24", font=("Segoe UI", 10, "bold"))
        self.lbl_odunc_kitap.pack(anchor="w", padx=12, pady=3)
        self.lbl_gecikmis_kitap = tk.Label(stats_frame, text="⚠️ Gecikmiş: 0", bg="#334155", fg="#f87171", font=("Segoe UI", 10, "bold"))
        self.lbl_gecikmis_kitap.pack(anchor="w", padx=12, pady=3)

        # Form alanları - Daha iyi aralık
        form_frame = tk.Frame(left_panel, bg=self.panel_color)
        form_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        self.entry_ad = self.create_input(form_frame, "📖 Kitap Adı")
        self.entry_yazar = self.create_input(form_frame, "✍️ Yazar")
        self.entry_tur = self.create_input(form_frame, "📂 Tür")
        self.entry_sayfa = self.create_input(form_frame, "📄 Sayfa Sayısı")
        self.entry_raf = self.create_input(form_frame, "📍 Raf No")
        self.entry_adet = self.create_input(form_frame, "🔢 Adet")
        self.entry_adet.insert(0, "1")

        # Butonlar - Modern stil
        btn_frame = tk.Frame(left_panel, bg=self.panel_color)
        btn_frame.pack(fill=tk.X, padx=20, pady=20, side=tk.BOTTOM)
        
        if self.kullanici_tipi == "ogretmen":
            ttk.Button(btn_frame, text="⚡ ÖDÜNÇ VER / İADE AL", command=self.akilli_islem_yap, 
                      style="Islem.TButton", cursor="hand2").pack(fill=tk.X, pady=4)
        
        ttk.Button(btn_frame, text="➕ KİTAP EKLE", command=self.kitap_ekle, 
                  style="Ekle.TButton", cursor="hand2").pack(fill=tk.X, pady=4)
        
        if self.kullanici_tipi == "ogretmen":
            ttk.Button(btn_frame, text="🗑️ SİL", command=self.kitap_sil, 
                      style="Sil.TButton", cursor="hand2").pack(fill=tk.X, pady=4)
        
        ttk.Button(btn_frame, text="🔄 TEMİZLE", command=self.formu_temizle, 
                  style="Normal.TButton", cursor="hand2").pack(fill=tk.X, pady=4)

        # SAĞ PANEL - Modern içerik alanı
        right_panel = tk.Frame(self.root, bg=self.bg_color)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # Üst arama çubuğu - Modern kart görünümü
        top_bar = tk.Frame(right_panel, bg=self.card_bg, padx=15, pady=12)
        top_bar.pack(fill=tk.X, padx=15, pady=(15, 0))
        
        # Arama
        search_frame = tk.Frame(top_bar, bg=self.card_bg)
        search_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        tk.Label(search_frame, text="🔍", bg=self.card_bg, font=("Arial", 14)).pack(side=tk.LEFT)
        self.entry_ara = ttk.Entry(search_frame, font=("Segoe UI", 11), width=30)
        self.entry_ara.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)
        self.entry_ara.bind("<KeyRelease>", self.arama_yap)
        
        # Sıralama
        sort_frame = tk.Frame(top_bar, bg=self.card_bg)
        sort_frame.pack(side=tk.RIGHT)
        
        self.sort_var = tk.StringVar(value="Yeniden Eskiye")
        self.combo_sort = ttk.Combobox(sort_frame, textvariable=self.sort_var, state="readonly", width=18, font=("Segoe UI", 10))
        self.combo_sort['values'] = ("Yeniden Eskiye", "Eskiden Yeniye", "Kitap Adı (A-Z)", "Yazar Adı (A-Z)")
        self.combo_sort.pack(side=tk.LEFT, padx=5)
        self.combo_sort.bind("<<ComboboxSelected>>", lambda e: self.verileri_guncelle())
        
        # Filtre butonları
        ttk.Button(sort_frame, text="📚 Tümü", command=lambda: self.verileri_guncelle(), 
                  style="Normal.TButton").pack(side=tk.LEFT, padx=3)
        ttk.Button(sort_frame, text="📖 Ödünçte", command=lambda: self.filtrele("Ödünç"), 
                  style="Normal.TButton").pack(side=tk.LEFT, padx=3)

        # Tablo alanı - Modern kart
        tree_frame = tk.Frame(right_panel, bg=self.card_bg, padx=5, pady=5)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        columns = ("ID", "Ad", "Yazar", "Tur", "Sayfa", "Raf", "Durum", "IadeTarihi")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="extended")
        headers = ["#", "Kitap Adı", "Yazar", "Tür", "Sayfa", "Raf", "Durum", "Son İade"]
        widths = [40, 280, 160, 100, 60, 60, 90, 100]
        
        for col, h, w in zip(columns, headers, widths):
            self.tree.heading(col, text=h)
            self.tree.column(col, width=w, anchor="center" if col != "Ad" else "w")
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Modern satır renkleri
        self.tree.tag_configure('odd', background='#f8fafc')
        self.tree.tag_configure('even', background='#ffffff')
        self.tree.tag_configure('odunc_normal', background='#fef3c7', foreground='#92400e')
        self.tree.tag_configure('odunc_yaklasan', background='#fef9c3', foreground='#a16207')
        self.tree.tag_configure('odunc_gecikmis', background='#fee2e2', foreground='#b91c1c')
        
        self.tree.bind("<Button-3>", self.sag_tik_goster)
        self.tree.bind("<Double-1>", lambda e: self.akilli_islem_yap())
        
        # Modern durum çubuğu
        self.status_bar = tk.Label(self.root, text="✓ Sistem Hazır", bg="#f1f5f9", fg="#64748b", 
                                   anchor="w", padx=15, pady=8, font=("Segoe UI", 9))
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def create_input(self, parent, title):
        """Modern input alanı oluştur"""
        tk.Label(parent, text=title, bg=self.panel_color, fg="#94a3b8", anchor="w", 
                 font=("Segoe UI", 9)).pack(fill=tk.X, pady=(12, 4))
        entry = ttk.Entry(parent, font=("Segoe UI", 10))
        entry.pack(fill=tk.X, ipady=6)
        return entry

    def durum_yaz(self, mesaj):
        self.status_bar.config(text=f"ℹ️ {mesaj}")
        self.root.after(4000, lambda: self.status_bar.config(text="✓ Sistem Hazır"))

    # --- WRAPPED ÖZELLİĞİ ---
    def wrapped_penceresi(self):
        top = tk.Toplevel(self.root)
        top.title("ŞAİK Wrapped - Yıllık Özet")
        top.geometry("800x600")
        top.configure(bg="#2C3E50")

        # Başlık ve Yıl Seçimi
        header = tk.Frame(top, bg="#2C3E50")
        header.pack(pady=20)
        tk.Label(header, text="🏆 YILLIK KÜTÜPHANE ÖZETİ", font=("Segoe UI", 24, "bold"), bg="#2C3E50", fg="#F1C40F").pack()
        
        yil_frame = tk.Frame(top, bg="#2C3E50")
        yil_frame.pack(pady=10)
        tk.Label(yil_frame, text="Yıl Seçiniz:", fg="white", bg="#2C3E50", font=("Segoe UI", 12)).pack(side=tk.LEFT, padx=10)
        
        current_year = str(datetime.now().year)
        yil_combo = ttk.Combobox(yil_frame, values=[str(y) for y in range(2024, 2030)], width=10, state="readonly")
        yil_combo.set(current_year)
        yil_combo.pack(side=tk.LEFT)

        content_frame = tk.Frame(top, bg="#2C3E50")
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        def istatistikleri_getir():
            yil = yil_combo.get()
            # Öncekileri temizle
            for widget in content_frame.winfo_children(): widget.destroy()

            # Verileri Çek (Hem aktif ödünçler hem geçmiş)
            veriler = []
            
            # 1. Aktif Ödünçlerden
            self.cursor.execute("SELECT k.ad, k.yazar, k.tur, o.ogrenci_ad, o.alinma_tarihi FROM odunc_alanlar o JOIN kitaplar k ON o.kitap_id = k.id")
            for row in self.cursor.fetchall():
                if row[4] and row[4].endswith(yil): veriler.append(row)
            
            # 2. Geçmişten
            self.cursor.execute("SELECT kitap_ad, yazar, tur, ogrenci_ad, alinma_tarihi FROM odunc_gecmisi")
            for row in self.cursor.fetchall():
                if row[4] and row[4].endswith(yil): veriler.append(row)

            if not veriler:
                tk.Label(content_frame, text=f"{yil} yılına ait veri bulunamadı.", bg="#2C3E50", fg="white", font=("Segoe UI", 14)).pack(pady=50)
                return

            # Hesaplamalar
            kitaplar = [v[0] for v in veriler]
            yazarlar = [v[1] for v in veriler]
            turler = [v[2] for v in veriler]
            ogrenciler = [v[3] for v in veriler]

            top_kitap = Counter(kitaplar).most_common(1)[0]
            top_yazar = Counter(yazarlar).most_common(1)[0]
            top_tur = Counter(turler).most_common(1)[0]
            top_ogrenci = Counter(ogrenciler).most_common(1)[0]
            toplam_okunan = len(veriler)

            # Kartları Oluştur
            self.kart_olustur(content_frame, "📚 YILIN KİTABI", f"{top_kitap[0]}\n({top_kitap[1]} kez okundu)", "#E74C3C", 0, 0)
            self.kart_olustur(content_frame, "✍️ YILIN YAZARI", f"{top_yazar[0]}\n({top_yazar[1]} kitap)", "#8E44AD", 0, 1)
            self.kart_olustur(content_frame, "🎭 EN SEVİLEN TÜR", f"{top_tur[0]}", "#2980B9", 1, 0)
            self.kart_olustur(content_frame, "🎓 KİTAP KURDU", f"{top_ogrenci[0]}\n({top_ogrenci[1]} kitap okudu)", "#F1C40F", 1, 1)
            
            tk.Label(content_frame, text=f"Bu yıl toplam {toplam_okunan} kitap ödünç verildi!", font=("Segoe UI", 12, "italic"), bg="#2C3E50", fg="#BDC3C7").grid(row=2, column=0, columnspan=2, pady=20)

        ttk.Button(yil_frame, text="GÖSTER", command=istatistikleri_getir, style="Islem.TButton").pack(side=tk.LEFT, padx=10)
        
        # İlk açılışta verileri getir
        istatistikleri_getir()

    def kart_olustur(self, parent, baslik, icerik, renk, r, c):
        frame = tk.Frame(parent, bg=renk, padx=5, pady=5)
        frame.grid(row=r, column=c, padx=10, pady=10, sticky="nsew")
        parent.grid_columnconfigure(c, weight=1)
        parent.grid_rowconfigure(r, weight=1)
        
        tk.Label(frame, text=baslik, bg=renk, fg="white", font=("Segoe UI", 12, "bold")).pack(pady=(10, 5))
        tk.Label(frame, text=icerik, bg=renk, fg="white", font=("Segoe UI", 14), wraplength=300).pack(pady=10)

    # --- YEDEKLEME ---
    def yedekle(self):
        try:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            src_file = os.path.join(base_dir, self.db_adi)
            hedef_dosya = filedialog.asksaveasfilename(defaultextension=".db", filetypes=[("Veritabanı Dosyası", "*.db")], initialfile=f"Yedek_{datetime.now().strftime('%Y%m%d')}_{self.db_adi}", title="Yedeği Kaydet")
            if hedef_dosya:
                shutil.copy2(src_file, hedef_dosya)
                messagebox.showinfo("Başarılı", f"Yedekleme tamamlandı:\n{hedef_dosya}")
        except Exception as e: messagebox.showerror("Hata", f"Yedekleme hatası: {e}")

    def ayarlar_penceresi(self):
        top = tk.Toplevel(self.root)
        top.title("Sistem Ayarları")
        top.geometry("300x200")
        top.configure(bg=self.bg_color)
        self.cursor.execute("SELECT deger FROM ayarlar WHERE anahtar='odunc_suresi'")
        mevcut_sure = self.cursor.fetchone()[0]
        tk.Label(top, text="Ödünç Verme Süresi (Gün)", bg=self.bg_color, font=("Segoe UI", 10, "bold")).pack(pady=20)
        entry_sure = ttk.Entry(top, font=("Segoe UI", 12), justify='center')
        entry_sure.insert(0, mevcut_sure)
        entry_sure.pack(pady=5, padx=20)
        def kaydet():
            yeni_sure = entry_sure.get()
            if not yeni_sure.isdigit():
                messagebox.showerror("Hata", "Lütfen geçerli bir sayı girin.", parent=top)
                return
            self.cursor.execute("UPDATE ayarlar SET deger=? WHERE anahtar='odunc_suresi'", (yeni_sure,))
            self.conn.commit()
            messagebox.showinfo("Başarılı", "Ayarlar kaydedildi. Yeni işlemler bu süreye göre yapılacak.", parent=top)
            top.destroy()
        ttk.Button(top, text="KAYDET", command=kaydet, style="Islem.TButton").pack(pady=20, fill=tk.X, padx=20)

    def verileri_guncelle(self):
        self.tree.delete(*self.tree.get_children())
        query = """SELECT k.id, k.ad, k.yazar, k.tur, k.sayfa_sayisi, k.raf_no, k.durum, o.iade_tarihi FROM kitaplar k LEFT JOIN odunc_alanlar o ON k.id = o.kitap_id"""
        self.cursor.execute(query)
        rows = self.cursor.fetchall()
        sort_option = self.combo_sort.get()
        if sort_option == "Ekleme Sırası (Yeniden Eskiye)": rows.sort(key=lambda x: x[0], reverse=True) 
        elif sort_option == "Ekleme Sırası (Eskiden Yeniye)": rows.sort(key=lambda x: x[0], reverse=False) 
        elif sort_option == "Kitap Adı (A-Z)": rows.sort(key=lambda x: self.tr_upper(x[1])) 
        elif sort_option == "Yazar Adı (A-Z)": rows.sort(key=lambda x: self.tr_upper(x[2])) 
        odunc_sayisi = 0
        gecikmis_sayisi = 0
        bugun = datetime.now()
        for i, row in enumerate(rows):
            durum = row[6]
            iade_tarihi_str = row[7]
            tag = 'even' if i % 2 == 0 else 'odd'
            if durum != 'Mevcut':
                odunc_sayisi += 1
                if iade_tarihi_str:
                    try:
                        iade_tarihi = datetime.strptime(iade_tarihi_str, "%d.%m.%Y")
                        kalan_gun = (iade_tarihi - bugun).days
                        if kalan_gun < 0:
                            tag = 'odunc_gecikmis'
                            gecikmis_sayisi += 1
                        elif kalan_gun <= 3: tag = 'odunc_yaklasan'
                        else: tag = 'odunc_normal'
                    except: tag = 'odunc_normal'
                else: tag = 'odunc_normal'
            display_row = list(row)
            if display_row[7] is None: display_row[7] = "-"
            self.tree.insert("", tk.END, values=display_row, tags=(tag,))
        self.lbl_toplam_kitap.config(text=f"📚 Toplam: {len(rows)}")
        self.lbl_odunc_kitap.config(text=f"📖 Ödünçte: {odunc_sayisi}")
        self.lbl_gecikmis_kitap.config(text=f"⚠️ Gecikmiş: {gecikmis_sayisi}")

    def akilli_islem_yap(self):
        secili = self.tree.selection()
        if not secili:
            messagebox.showwarning("Seçim Yok", "Lütfen listeden bir kitap seçiniz.")
            return
        item = self.tree.item(secili)
        durum = item['values'][6]
        if durum == 'Mevcut': self.odunc_ver_penceresi()
        else:
            popup = tk.Toplevel(self.root)
            popup.title("İşlem Seç")
            popup.geometry("300x150")
            popup.configure(bg=self.bg_color)
            tk.Label(popup, text=f"Seçili Kitap: {item['values'][1]}", bg=self.bg_color, font=("Segoe UI", 10, "bold")).pack(pady=10)
            ttk.Button(popup, text="ℹ️ KİMDE? (Bilgi Göster)", command=lambda: [self.odunc_bilgisi_goster(), popup.destroy()], style="Normal.TButton").pack(fill=tk.X, padx=20, pady=5)
            ttk.Button(popup, text="✅ İADE AL (Rafa Kaldır)", command=lambda: [self.iade_al(), popup.destroy()], style="Islem.TButton").pack(fill=tk.X, padx=20, pady=5)

    def kitap_ekle(self):
        adet_str = self.entry_adet.get().strip()
        adet = int(adet_str) if adet_str.isdigit() and int(adet_str) > 0 else 1
        
        veriler = (self.entry_ad.get(), self.entry_yazar.get(), self.entry_tur.get(), 
                   self.entry_sayfa.get(), self.entry_raf.get(), adet)
        if not veriler[0] or not veriler[1]:
            messagebox.showwarning("Eksik Bilgi", "Kitap Adı ve Yazar alanları zorunludur.")
            return
        try:
            self.cursor.execute("INSERT INTO kitaplar (ad, yazar, tur, sayfa_sayisi, raf_no, adet) VALUES (?,?,?,?,?,?)", veriler)
            self.conn.commit()
            self.verileri_guncelle()
            self.formu_temizle()
            self.durum_yaz(f"Kitap eklendi ({adet} adet).")
        except Exception as e: messagebox.showerror("Hata", str(e))

    def kitap_sil(self):
        secili = self.tree.selection()
        if secili:
            item = self.tree.item(secili)
            if messagebox.askyesno("Sil", f"'{item['values'][1]}' kitabını silmek istediğinize emin misiniz?"):
                kitap_id = item['values'][0]
                
                # Barkod dosyasını sil
                self.cursor.execute("SELECT barkod FROM kitaplar WHERE id=?", (kitap_id,))
                barkod_result = self.cursor.fetchone()
                if barkod_result and barkod_result[0]:
                    barkod_kodu = barkod_result[0]
                    base_dir = os.path.dirname(os.path.abspath(__file__))
                    barkod_dosya = os.path.join(base_dir, "barkodlar", f"{barkod_kodu}.png")
                    if os.path.exists(barkod_dosya):
                        try:
                            os.remove(barkod_dosya)
                        except:
                            pass
                
                # Kitabı ve ödünç kayıtlarını sil
                self.cursor.execute("DELETE FROM kitaplar WHERE id=?", (kitap_id,))
                self.cursor.execute("DELETE FROM odunc_alanlar WHERE kitap_id=?", (kitap_id,))
                self.conn.commit()
                self.verileri_guncelle()
                self.durum_yaz("Kitap ve barkodu silindi.")

    def odunc_ver_penceresi(self):
        secili = self.tree.selection()
        item = self.tree.item(secili)
        kitap_id = item['values'][0]
        top = tk.Toplevel(self.root)
        top.title("Ödünç Verme İşlemi")
        top.geometry("400x400")
        top.configure(bg=self.bg_color)
        tk.Label(top, text="Öğrenci Bilgileri", font=("Segoe UI", 14, "bold"), bg=self.bg_color, fg=self.panel_color).pack(pady=20)
        entries = {}
        for alan in ["Öğrenci Adı Soyadı", "Okul No", "Sınıf"]:
            frame = tk.Frame(top, bg=self.bg_color)
            frame.pack(fill=tk.X, padx=30, pady=5)
            tk.Label(frame, text=alan, bg=self.bg_color, width=15, anchor="w").pack(side=tk.LEFT)
            e = ttk.Entry(frame)
            e.pack(side=tk.RIGHT, fill=tk.X, expand=True)
            entries[alan] = e
        def onayla():
            if not all(e.get() for e in entries.values()):
                messagebox.showwarning("Eksik", "Tüm alanları doldurunuz.", parent=top)
                return
            self.cursor.execute("SELECT deger FROM ayarlar WHERE anahtar='odunc_suresi'")
            gun_sayisi = int(self.cursor.fetchone()[0])
            bugun = datetime.now()
            iade = bugun + timedelta(days=gun_sayisi)
            tarih_fmt = "%d.%m.%Y"
            self.cursor.execute("UPDATE kitaplar SET durum='Ödünç Verildi' WHERE id=?", (kitap_id,))
            self.cursor.execute("INSERT INTO odunc_alanlar (kitap_id, ogrenci_ad, ogrenci_no, sinif, alinma_tarihi, iade_tarihi) VALUES (?,?,?,?,?,?)", 
                                (kitap_id, entries["Öğrenci Adı Soyadı"].get(), entries["Okul No"].get(), entries["Sınıf"].get(), bugun.strftime(tarih_fmt), iade.strftime(tarih_fmt)))
            self.conn.commit()
            self.verileri_guncelle()
            top.destroy()
            self.durum_yaz(f"Kitap verildi. Son iade tarihi: {iade.strftime(tarih_fmt)}")
            messagebox.showinfo("Başarılı", f"İşlem Tamam!\nÖğrenciye '{iade.strftime(tarih_fmt)}' tarihine kadar süre verildi ({gun_sayisi} Gün).")
        ttk.Button(top, text="ONAYLA VE VER", command=onayla, style="Ekle.TButton").pack(fill=tk.X, padx=30, pady=30)

    def iade_al(self):
        secili = self.tree.selection()
        if not secili:
            messagebox.showwarning("Seçim Yok", "Lütfen iade alınacak kitabı seçin.")
            return
        
        # Tek kitap için
        if len(secili) == 1:
            item = self.tree.item(secili[0])
            kitap_id = item['values'][0]
            kitap_ad = item['values'][1]
            
            # Önce verileri al
            self.cursor.execute("""
                SELECT k.ad, k.yazar, k.tur, o.ogrenci_ad, o.ogrenci_no, o.sinif, o.alinma_tarihi, o.iade_tarihi
                FROM odunc_alanlar o 
                JOIN kitaplar k ON o.kitap_id = k.id 
                WHERE o.kitap_id = ?""", (kitap_id,))
            veri = self.cursor.fetchone()
            
            if not veri:
                messagebox.showwarning("Uyarı", "Bu kitap zaten iade edilmiş.")
                return
            
            # Yorum penceresi aç
            self.yorum_ile_iade(kitap_id, kitap_ad, veri)
        else:
            # Çoklu iade
            self.toplu_iade(secili)
    
    def yorum_ile_iade(self, kitap_id, kitap_ad, veri):
        """İade sırasında yorum alma penceresi"""
        top = tk.Toplevel(self.root)
        top.title("📖 Kitap İade ve Değerlendirme")
        top.geometry("450x400")
        top.configure(bg=self.bg_color)
        top.resizable(False, False)
        
        tk.Label(top, text="📖 Kitap Değerlendirmesi", font=("Segoe UI", 14, "bold"),
                 bg=self.bg_color, fg=self.panel_color).pack(pady=15)
        
        tk.Label(top, text=f"'{kitap_ad[:40]}...'", font=("Segoe UI", 11),
                 bg=self.bg_color, wraplength=400).pack()
        
        tk.Label(top, text=f"Öğrenci: {veri[3]}", font=("Segoe UI", 10),
                 bg=self.bg_color, fg="#666").pack(pady=5)
        
        # Puan seçimi
        puan_frame = tk.Frame(top, bg=self.bg_color)
        puan_frame.pack(pady=15)
        
        tk.Label(puan_frame, text="Puan:", bg=self.bg_color, font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=5)
        
        puan_var = tk.IntVar(value=5)
        for i in range(1, 6):
            emoji = "⭐" * i
            rb = tk.Radiobutton(puan_frame, text=emoji, variable=puan_var, value=i,
                                bg=self.bg_color, font=("Segoe UI", 10))
            rb.pack(side=tk.LEFT, padx=2)
        
        # Yorum alanı
        tk.Label(top, text="Kitap hakkında düşünceleriniz (isteğe bağlı):", 
                 bg=self.bg_color, font=("Segoe UI", 10)).pack(pady=(10, 5))
        
        yorum_text = tk.Text(top, height=4, width=45, font=("Segoe UI", 10))
        yorum_text.pack(padx=20, pady=5)
        yorum_text.insert("1.0", "")
        
        def tamamla():
            # Yorum kaydet
            yorum = yorum_text.get("1.0", tk.END).strip()
            puan = puan_var.get()
            
            if yorum:  # Sadece yorum varsa kaydet
                self.cursor.execute("""
                    INSERT INTO yorumlar (kitap_id, ogrenci_ad, ogrenci_no, yorum, puan, tarih)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (kitap_id, veri[3], veri[4], yorum, puan, datetime.now().strftime("%d.%m.%Y")))
            
            # Arşive kaydet
            bugun = datetime.now().strftime("%d.%m.%Y")
            self.cursor.execute("""
                INSERT INTO odunc_gecmisi (kitap_ad, yazar, tur, ogrenci_ad, ogrenci_no, sinif, alinma_tarihi, iade_tarihi)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (veri[0], veri[1], veri[2], veri[3], veri[4], veri[5], veri[6], bugun))
            
            # Puan hesapla
            try:
                alinma = datetime.strptime(veri[6], "%d.%m.%Y")
                gun_farki = (datetime.now() - alinma).days
                kazanilan_puan = 10
                if gun_farki <= 7:
                    kazanilan_puan += 5
                elif gun_farki <= 14:
                    kazanilan_puan += 3
                
                self.cursor.execute("""
                    UPDATE ogrenciler 
                    SET toplam_puan = toplam_puan + ?, toplam_kitap = toplam_kitap + 1
                    WHERE okul_no = ?
                """, (kazanilan_puan, veri[4]))
            except:
                pass
            
            # İade işlemini tamamla
            self.cursor.execute("UPDATE kitaplar SET durum='Mevcut' WHERE id=?", (kitap_id,))
            self.cursor.execute("DELETE FROM odunc_alanlar WHERE kitap_id=?", (kitap_id,))
            self.conn.commit()
            
            top.destroy()
            self.verileri_guncelle()
            
            if yorum:
                self.durum_yaz("Kitap iade alındı ve yorum kaydedildi.")
                messagebox.showinfo("Teşekkürler! 📚", "Değerlendirmeniz kaydedildi.\nDiğer öğrencilere yardımcı olacak!")
            else:
                self.durum_yaz("Kitap iade alındı.")
        
        def atla():
            # Yorum olmadan iade
            bugun = datetime.now().strftime("%d.%m.%Y")
            self.cursor.execute("""
                INSERT INTO odunc_gecmisi (kitap_ad, yazar, tur, ogrenci_ad, ogrenci_no, sinif, alinma_tarihi, iade_tarihi)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (veri[0], veri[1], veri[2], veri[3], veri[4], veri[5], veri[6], bugun))
            
            self.cursor.execute("UPDATE kitaplar SET durum='Mevcut' WHERE id=?", (kitap_id,))
            self.cursor.execute("DELETE FROM odunc_alanlar WHERE kitap_id=?", (kitap_id,))
            self.conn.commit()
            
            top.destroy()
            self.verileri_guncelle()
            self.durum_yaz("Kitap iade alındı.")
        
        btn_frame = tk.Frame(top, bg=self.bg_color)
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="✅ KAYDET VE İADE AL", command=tamamla,
                  font=("Segoe UI", 11, "bold"), bg=self.action_color, fg="white",
                  width=18, cursor="hand2").pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="⏭️ ATLA", command=atla,
                  font=("Segoe UI", 11), bg="#95A5A6", fg="white",
                  width=10, cursor="hand2").pack(side=tk.LEFT, padx=5)

    def rapor_al(self):
        try:
            dosya_yolu = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV Dosyası", "*.csv"), ("Tüm Dosyalar", "*.*")], title="Raporu Kaydet")
            if not dosya_yolu: return
            query = """SELECT k.id, k.ad, k.yazar, k.tur, k.sayfa_sayisi, k.raf_no, k.durum, o.ogrenci_ad, o.ogrenci_no, o.sinif, o.alinma_tarihi, o.iade_tarihi 
                       FROM kitaplar k LEFT JOIN odunc_alanlar o ON k.id = o.kitap_id ORDER BY k.id ASC"""
            self.cursor.execute(query)
            rows = self.cursor.fetchall()
            with open(dosya_yolu, mode='w', newline='', encoding='utf-8-sig') as file:
                writer = csv.writer(file, delimiter=';')
                writer.writerow(["ID", "Kitap Adı", "Yazar", "Tür", "Sayfa", "Raf No", "Durum", "Öğrenci Adı", "Öğrenci No", "Sınıf", "Veriliş Tarihi", "Son İade Tarihi"])
                for row in rows: writer.writerow(row)
            messagebox.showinfo("Başarılı", f"Rapor kaydedildi:\n{dosya_yolu}")
        except Exception as e: messagebox.showerror("Hata", f"Rapor hatası: {e}")

    def odunc_bilgisi_goster(self):
        secili = self.tree.selection()
        kitap_id = self.tree.item(secili)['values'][0]
        self.cursor.execute("SELECT * FROM odunc_alanlar WHERE kitap_id=?", (kitap_id,))
        bilgi = self.cursor.fetchone()
        if bilgi: messagebox.showinfo("Teslim Bilgisi", f"Öğrenci: {bilgi[2]}\nNo: {bilgi[3]}\nSınıf: {bilgi[4]}\n\nVeriliş: {bilgi[5]}\nSon Tarih: {bilgi[6]}")

    def sag_tik_goster(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            # Eğer tıklanan satır seçili değilse, sadece onu seç
            if item not in self.tree.selection():
                self.tree.selection_set(item)
            
            secili_sayisi = len(self.tree.selection())
            durum = self.tree.item(item)['values'][6]
            self.context_menu.delete(0, tk.END)
            
            # Çoklu seçim varsa
            if secili_sayisi > 1:
                self.context_menu.add_command(label=f"📚 {secili_sayisi} kitap seçili", state="disabled")
                self.context_menu.add_separator()
                self.context_menu.add_command(label="➕ Toplu Ödünç Ver", command=self.toplu_odunc)
                self.context_menu.add_command(label="✅ Toplu İade Al", command=lambda: self.toplu_iade(self.tree.selection()))
                self.context_menu.add_separator()
                self.context_menu.add_command(label="❌ Seçimi Temizle", command=lambda: self.tree.selection_remove(*self.tree.selection()))
            else:
                # Tek kitap seçimi
                self.context_menu.add_command(label="✏️ Düzenle", command=self.kitap_duzenle_penceresi)
                self.context_menu.add_command(label="💬 Yorumları Gör", command=self.yorumlari_goster)
                self.context_menu.add_separator()
                
                if durum == 'Mevcut':
                    self.context_menu.add_command(label="➕ Ödünç Ver", command=self.odunc_ver_penceresi)
                    self.context_menu.add_command(label="📅 Rezerve Et", command=self.hizli_rezervasyon)
                else:
                    self.context_menu.add_command(label="ℹ️ Kimde? (Bilgi)", command=self.odunc_bilgisi_goster)
                    self.context_menu.add_command(label="✅ İade Al", command=self.iade_al)
                    self.context_menu.add_command(label="📅 Sıraya Gir", command=self.hizli_rezervasyon)
                
                self.context_menu.add_separator()
                if self.kullanici_tipi == "ogretmen":
                    self.context_menu.add_command(label="🗑️ Sil", command=self.kitap_sil)
            
            self.context_menu.post(event.x_root, event.y_root)

    def formu_temizle(self):
        for e in [self.entry_ad, self.entry_yazar, self.entry_tur, self.entry_sayfa, self.entry_raf, self.entry_adet]: 
            e.delete(0, tk.END)
        self.entry_adet.insert(0, "1")
        self.durum_yaz("Form temizlendi.")

    def filtrele(self, mod):
        self.tree.delete(*self.tree.get_children())
        if mod == "Ödünç": query = "SELECT k.id, k.ad, k.yazar, k.tur, k.sayfa_sayisi, k.raf_no, k.durum, o.iade_tarihi FROM kitaplar k LEFT JOIN odunc_alanlar o ON k.id = o.kitap_id WHERE k.durum != 'Mevcut' ORDER BY k.id DESC"
        else: query = "SELECT k.id, k.ad, k.yazar, k.tur, k.sayfa_sayisi, k.raf_no, k.durum, o.iade_tarihi FROM kitaplar k LEFT JOIN odunc_alanlar o ON k.id = o.kitap_id ORDER BY k.id DESC"
        self.cursor.execute(query)
        rows = self.cursor.fetchall()
        bugun = datetime.now()
        for i, row in enumerate(rows):
            durum = row[6]
            iade_tarihi_str = row[7]
            tag = 'odunc_normal'
            if durum != 'Mevcut' and iade_tarihi_str:
                 try:
                    iade_tarihi = datetime.strptime(iade_tarihi_str, "%d.%m.%Y")
                    kalan_gun = (iade_tarihi - bugun).days
                    if kalan_gun < 0: tag = 'odunc_gecikmis'
                    elif kalan_gun <= 3: tag = 'odunc_yaklasan'
                 except: pass
            display_row = list(row)
            if display_row[7] is None: display_row[7] = "-"
            self.tree.insert("", tk.END, values=display_row, tags=(tag,))

    def arama_yap(self, event):
        anahtar = self.entry_ara.get()
        self.tree.delete(*self.tree.get_children())
        query = "SELECT k.id, k.ad, k.yazar, k.tur, k.sayfa_sayisi, k.raf_no, k.durum, o.iade_tarihi FROM kitaplar k LEFT JOIN odunc_alanlar o ON k.id = o.kitap_id WHERE k.ad LIKE ? OR k.yazar LIKE ?"
        self.cursor.execute(query, (f"%{anahtar}%", f"%{anahtar}%"))
        rows = self.cursor.fetchall()
        for row in rows:
            tag = 'even'
            if row[6] != 'Mevcut': tag = 'odunc_normal'
            display_row = list(row)
            if display_row[7] is None: display_row[7] = "-"
            self.tree.insert("", tk.END, values=display_row, tags=(tag,))

    # --- EXCEL IMPORT ---
    def excel_import(self):
        """Excel dosyasından kitap aktarma"""
        if not EXCEL_DESTEGI:
            messagebox.showerror("Hata", "Excel desteği için 'openpyxl' kütüphanesini yükleyin:\npip install openpyxl")
            return
        
        dosya = filedialog.askopenfilename(
            title="Excel Dosyası Seç",
            filetypes=[("Excel Dosyaları", "*.xlsx *.xls"), ("Tüm Dosyalar", "*.*")]
        )
        
        if not dosya:
            return
        
        try:
            wb = load_workbook(dosya)
            ws = wb.active
            
            # Başlıkları bul
            basliklar = []
            for cell in ws[1]:
                basliklar.append(str(cell.value).lower().strip() if cell.value else "")
            
            # Sütun eşleştirme
            sutun_map = {
                'ad': None, 'yazar': None, 'tur': None, 
                'sayfa': None, 'raf': None, 'adet': None
            }
            
            for i, baslik in enumerate(basliklar):
                if 'kitap' in baslik and 'ad' in baslik:
                    sutun_map['ad'] = i
                elif 'ad' in baslik and sutun_map['ad'] is None:
                    sutun_map['ad'] = i
                elif 'yazar' in baslik:
                    sutun_map['yazar'] = i
                elif 'tür' in baslik or 'tur' in baslik:
                    sutun_map['tur'] = i
                elif 'sayfa' in baslik:
                    sutun_map['sayfa'] = i
                elif 'raf' in baslik:
                    sutun_map['raf'] = i
                elif 'adet' in baslik or 'miktar' in baslik or 'sayı' in baslik:
                    sutun_map['adet'] = i
            
            # Verileri oku
            kitaplar = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                
                # Adet değerini al
                adet_val = 1
                if sutun_map['adet'] is not None and row[sutun_map['adet']]:
                    try:
                        adet_val = int(row[sutun_map['adet']])
                        if adet_val < 1: adet_val = 1
                    except: adet_val = 1
                
                kitap = {
                    'ad': str(row[sutun_map['ad']]) if sutun_map['ad'] is not None and row[sutun_map['ad']] else "",
                    'yazar': str(row[sutun_map['yazar']]) if sutun_map['yazar'] is not None and row[sutun_map['yazar']] else "",
                    'tur': str(row[sutun_map['tur']]) if sutun_map['tur'] is not None and row[sutun_map['tur']] else "",
                    'sayfa': row[sutun_map['sayfa']] if sutun_map['sayfa'] is not None else None,
                    'raf': str(row[sutun_map['raf']]) if sutun_map['raf'] is not None and row[sutun_map['raf']] else "",
                    'adet': adet_val
                }
                
                if kitap['ad']:
                    kitaplar.append(kitap)
            
            if not kitaplar:
                messagebox.showwarning("Uyarı", "Excel dosyasında geçerli kitap bulunamadı.")
                return
            
            # Onay iste
            onay = messagebox.askyesno(
                "Onay", 
                f"{len(kitaplar)} kitap bulundu.\n\nBu kitapları veritabanına eklemek istiyor musunuz?"
            )
            
            if onay:
                eklenen = 0
                for kitap in kitaplar:
                    try:
                        self.cursor.execute(
                            "INSERT INTO kitaplar (ad, yazar, tur, sayfa_sayisi, raf_no, adet) VALUES (?,?,?,?,?,?)",
                            (kitap['ad'], kitap['yazar'], kitap['tur'], kitap['sayfa'], kitap['raf'], kitap['adet'])
                        )
                        eklenen += 1
                    except Exception as e:
                        print(f"Kitap eklenemedi: {kitap['ad']} - {e}")
                
                self.conn.commit()
                self.verileri_guncelle()
                messagebox.showinfo("Başarılı", f"{eklenen} kitap başarıyla eklendi.")
                self.durum_yaz(f"Excel'den {eklenen} kitap aktarıldı.")
                
        except Exception as e:
            messagebox.showerror("Hata", f"Excel dosyası okunamadı:\n{e}")
    
    # --- BARKOD OLUŞTURMA ---
    def barkod_olustur(self):
        """Seçili kitap için barkod oluştur"""
        if not BARKOD_OLUSTURMA_DESTEGI:
            messagebox.showerror("Hata", "Barkod desteği için 'python-barcode' ve 'pillow' yükleyin:\npip install python-barcode pillow")
            return
        
        secili = self.tree.selection()
        if not secili:
            messagebox.showwarning("Seçim Yok", "Lütfen listeden bir kitap seçiniz.")
            return
        
        item = self.tree.item(secili)
        kitap_id = item['values'][0]
        kitap_ad = item['values'][1]
        
        # Mevcut barkodu kontrol et
        self.cursor.execute("SELECT barkod FROM kitaplar WHERE id=?", (kitap_id,))
        mevcut = self.cursor.fetchone()[0]
        
        if mevcut:
            # Mevcut barkodu göster (her kitap için 1 barkod)
            self.barkod_goster(kitap_id)
            return
        
        # Benzersiz barkod oluştur (kitap ID'si ile)
        barkod_kodu = self.benzersiz_barkod_olustur(kitap_id)
        
        # Veritabanını güncelle
        self.cursor.execute("UPDATE kitaplar SET barkod=? WHERE id=?", (barkod_kodu, kitap_id))
        self.conn.commit()
        
        # Barkod görselini kaydet
        self.barkod_kaydet(barkod_kodu, kitap_ad)
        
        self.durum_yaz(f"Barkod oluşturuldu: {barkod_kodu}")
        self.barkod_goster(kitap_id)
    
    def benzersiz_barkod_olustur(self, kitap_id=None):
        """Benzersiz barkod kodu oluştur - Code128 formatı"""
        while True:
            # Basit format: SAIK + ID veya rastgele sayı
            if kitap_id:
                kod = f"SAIK{kitap_id:05d}"
            else:
                kod = f"SAIK{random.randint(10000, 99999)}"
            
            self.cursor.execute("SELECT id FROM kitaplar WHERE barkod=?", (kod,))
            if not self.cursor.fetchone():
                return kod
    
    def barkod_kaydet(self, barkod_kodu, kitap_ad):
        """Barkod görselini PNG olarak kaydet - Code128 formatı"""
        try:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            barkod_dir = os.path.join(base_dir, "barkodlar")
            
            if not os.path.exists(barkod_dir):
                os.makedirs(barkod_dir)
            
            # Code128 barkod oluştur (daha esnek, her karakter destekler)
            CODE128 = barcode.get_barcode_class('code128')
            code = CODE128(barkod_kodu, writer=ImageWriter())
            
            # Dosya yolu
            dosya_yolu = os.path.join(barkod_dir, barkod_kodu)
            
            code.save(dosya_yolu)
            return dosya_yolu + ".png"
        except Exception as e:
            print(f"Barkod kaydetme hatası: {e}")
            return None
    
    def barkod_goster(self, kitap_id):
        """Barkod görüntüleme ve yazdırma penceresi"""
        self.cursor.execute("SELECT ad, barkod FROM kitaplar WHERE id=?", (kitap_id,))
        sonuc = self.cursor.fetchone()
        
        if not sonuc or not sonuc[1]:
            messagebox.showwarning("Uyarı", "Bu kitabın barkodu bulunamadı.")
            return
        
        kitap_ad, barkod_kodu = sonuc
        
        top = tk.Toplevel(self.root)
        top.title(f"📊 Barkod - {barkod_kodu}")
        top.geometry("450x420")
        top.configure(bg="white")
        top.resizable(False, False)
        
        # Başlık
        tk.Label(top, text="📊 Kitap Barkodu", font=("Segoe UI", 14, "bold"),
                 bg="white", fg="#1e293b").pack(pady=(15, 5))
        
        # Kitap adı
        tk.Label(top, text=kitap_ad, font=("Segoe UI", 11),
                 bg="white", fg="#64748b", wraplength=400).pack(pady=5)
        
        # Barkod görselini bul/oluştur
        base_dir = os.path.dirname(os.path.abspath(__file__))
        barkod_dir = os.path.join(base_dir, "barkodlar")
        barkod_dosya = os.path.join(barkod_dir, f"{barkod_kodu}.png")
        
        # Eğer dosya yoksa oluştur
        if not os.path.exists(barkod_dosya):
            self.barkod_kaydet(barkod_kodu, kitap_ad)
        
        # Barkod görselini göster
        barkod_frame = tk.Frame(top, bg="white", padx=20, pady=15)
        barkod_frame.pack(fill=tk.X)
        
        if os.path.exists(barkod_dosya) and PIL_DESTEGI:
            try:
                img = Image.open(barkod_dosya)
                # Orantılı boyutlandır
                w, h = img.size
                new_w = 380
                new_h = int(h * new_w / w)
                img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                label = tk.Label(barkod_frame, image=photo, bg="white")
                label.image = photo
                label.pack()
            except Exception as e:
                tk.Label(barkod_frame, text="[Barkod yüklenemedi]", bg="white", fg="gray",
                         font=("Segoe UI", 12)).pack(pady=20)
        else:
            tk.Label(barkod_frame, text="[Barkod görseli yok]", bg="white", fg="gray",
                     font=("Segoe UI", 12)).pack(pady=20)
        
        # Barkod numarası
        tk.Label(top, text=barkod_kodu, font=("Consolas", 16, "bold"),
                 bg="white", fg="#1e293b").pack(pady=10)
        
        # Butonlar
        btn_frame = tk.Frame(top, bg="white")
        btn_frame.pack(pady=15)
        
        def kopyala():
            self.root.clipboard_clear()
            self.root.clipboard_append(barkod_kodu)
            messagebox.showinfo("✓ Kopyalandı", f"Barkod panoya kopyalandı:\n{barkod_kodu}")
        
        def yazdir():
            if os.path.exists(barkod_dosya):
                try:
                    # macOS için
                    if sys.platform == 'darwin':
                        os.system(f'open -a Preview "{barkod_dosya}"')
                    # Windows için
                    elif sys.platform == 'win32':
                        os.startfile(barkod_dosya, 'print')
                    else:
                        messagebox.showinfo("Yazdır", f"Barkod dosyası:\n{barkod_dosya}")
                except Exception as e:
                    messagebox.showerror("Hata", f"Yazdırma hatası: {e}")
            else:
                messagebox.showwarning("Uyarı", "Barkod dosyası bulunamadı.")
        
        tk.Button(btn_frame, text="📋 Kopyala", command=kopyala,
                  font=("Segoe UI", 10), bg="#3b82f6", fg="white",
                  width=12, cursor="hand2").pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="🖨️ Yazdır", command=yazdir,
                  font=("Segoe UI", 10), bg="#10b981", fg="white",
                  width=12, cursor="hand2").pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="✕ Kapat", command=top.destroy,
                  font=("Segoe UI", 10), bg="#64748b", fg="white",
                  width=12, cursor="hand2").pack(side=tk.LEFT, padx=5)
    
    def toplu_barkod_olustur(self):
        """Barkodu olmayan tüm kitaplara barkod oluştur"""
        self.cursor.execute("SELECT id, ad FROM kitaplar WHERE barkod IS NULL OR barkod = ''")
        kitaplar = self.cursor.fetchall()
        
        if not kitaplar:
            messagebox.showinfo("Bilgi", "Tüm kitapların barkodu zaten mevcut.")
            return
        
        onay = messagebox.askyesno("Onay", 
            f"{len(kitaplar)} kitaba barkod oluşturulacak.\n\nDevam etmek istiyor musunuz?")
        
        if not onay:
            return
        
        olusturulan = 0
        for kitap_id, kitap_ad in kitaplar:
            try:
                barkod_kodu = self.benzersiz_barkod_olustur()
                self.cursor.execute("UPDATE kitaplar SET barkod=? WHERE id=?", (barkod_kodu, kitap_id))
                self.barkod_kaydet(barkod_kodu, kitap_ad)
                olusturulan += 1
            except Exception as e:
                print(f"Barkod oluşturulamadı: {kitap_ad} - {e}")
        
        self.conn.commit()
        messagebox.showinfo("Başarılı", f"{olusturulan} kitap için barkod oluşturuldu.")
        self.durum_yaz(f"{olusturulan} kitaba barkod eklendi.")
    
    # --- BARKOD OKUMA ---
    def barkod_tara(self):
        """Kamera ile barkod tara"""
        if not BARKOD_OKUMA_DESTEGI:
            messagebox.showerror("Hata", "Barkod tarama için gerekli kütüphaneleri yükleyin:\npip install opencv-python pyzbar\n\nAyrıca macOS için: brew install zbar")
            return
        
        # Kamera erişimini kontrol et
        try:
            cap = cv2.VideoCapture(0)
            if not cap.isOpened():
                messagebox.showerror("Kamera Hatası", "Kamera açılamadı!\n\nLütfen kamera izinlerini kontrol edin:\nSistem Tercihleri → Gizlilik ve Güvenlik → Kamera")
                return
        except Exception as e:
            messagebox.showerror("Kamera Hatası", f"Kamera başlatılamadı:\n{e}")
            return
        
        # Kamera penceresi
        tarama_pencere = tk.Toplevel(self.root)
        tarama_pencere.title("Barkod Tarama")
        tarama_pencere.geometry("700x580")
        tarama_pencere.configure(bg="#2C3E50")
        
        tk.Label(tarama_pencere, text="📷 Barkodu kameraya gösterin", 
                 font=("Segoe UI", 14, "bold"), bg="#2C3E50", fg="white").pack(pady=10)
        
        video_label = tk.Label(tarama_pencere, bg="black", width=640, height=400)
        video_label.pack(pady=10)
        
        sonuc_label = tk.Label(tarama_pencere, text="🔍 Barkod bekleniyor...", 
                                font=("Segoe UI", 12), bg="#2C3E50", fg="#BDC3C7")
        sonuc_label.pack(pady=10)
        
        running = [True]
        
        def update_frame():
            if not running[0]:
                return
            
            try:
                ret, frame = cap.read()
                if not ret or frame is None:
                    if running[0]:
                        tarama_pencere.after(100, update_frame)
                    return
                
                # Barkod algıla
                try:
                    barcodes = pyzbar.decode(frame)
                except Exception as e:
                    print(f"Barkod decode hatası: {e}")
                    barcodes = []
                
                for barcode_obj in barcodes:
                    try:
                        barkod_data = barcode_obj.data.decode('utf-8')
                    except:
                        continue
                    
                    # Çerçeve çiz
                    (x, y, w, h) = barcode_obj.rect
                    cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 3)
                    cv2.putText(frame, barkod_data, (x, y - 10), 
                               cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                    
                    # Veritabanında ara
                    self.cursor.execute("SELECT id, ad FROM kitaplar WHERE barkod=?", (barkod_data,))
                    sonuc = self.cursor.fetchone()
                    
                    if sonuc:
                        sonuc_label.config(text=f"✅ Bulundu: {sonuc[1]}", fg="#2ECC71")
                        running[0] = False
                        cap.release()
                        
                        # Kitabı seç
                        for item in self.tree.get_children():
                            if self.tree.item(item)['values'][0] == sonuc[0]:
                                self.tree.selection_set(item)
                                self.tree.see(item)
                                break
                        
                        self.durum_yaz(f"Barkod ile bulundu: {sonuc[1]}")
                        tarama_pencere.after(1500, tarama_pencere.destroy)
                        return
                    else:
                        sonuc_label.config(text=f"❌ Sistemde bulunamadı: {barkod_data}", fg="#E74C3C")
                
                # Frame'i göster
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                frame = cv2.resize(frame, (640, 400))
                img = Image.fromarray(frame)
                imgtk = ImageTk.PhotoImage(image=img)
                video_label.imgtk = imgtk
                video_label.configure(image=imgtk)
                
            except Exception as e:
                print(f"Frame işleme hatası: {e}")
            
            if running[0]:
                tarama_pencere.after(30, update_frame)
        
        def kapat():
            running[0] = False
            try:
                cap.release()
            except:
                pass
            tarama_pencere.destroy()
        
        tarama_pencere.protocol("WM_DELETE_WINDOW", kapat)
        
        btn_frame = tk.Frame(tarama_pencere, bg="#2C3E50")
        btn_frame.pack(pady=10)
        
        tk.Button(btn_frame, text="❌ Kapat", command=kapat,
                  font=("Segoe UI", 11), bg="#C0392B", fg="white", width=12).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🔄 Yeniden Başlat", 
                  command=lambda: [kapat(), self.root.after(500, self.barkod_tara)],
                  font=("Segoe UI", 11), bg="#3498DB", fg="white", width=12).pack(side=tk.LEFT, padx=5)
        
        update_frame()
    
    def barkod_ile_ara(self):
        """Manuel barkod girişi ile arama"""
        top = tk.Toplevel(self.root)
        top.title("Barkod ile Ara")
        top.geometry("400x180")
        top.configure(bg=self.bg_color)
        
        tk.Label(top, text="Barkod Numarası:", font=("Segoe UI", 12), 
                 bg=self.bg_color).pack(pady=20)
        
        entry = ttk.Entry(top, font=("Segoe UI", 14), width=20, justify="center")
        entry.pack(pady=5)
        entry.focus_set()
        
        def ara(event=None):
            barkod = entry.get().strip()
            if not barkod:
                return
            
            self.cursor.execute("SELECT id, ad FROM kitaplar WHERE barkod=?", (barkod,))
            sonuc = self.cursor.fetchone()
            
            if sonuc:
                # Kitabı seç
                for item in self.tree.get_children():
                    if self.tree.item(item)['values'][0] == sonuc[0]:
                        self.tree.selection_set(item)
                        self.tree.see(item)
                        break
                top.destroy()
                self.durum_yaz(f"Bulundu: {sonuc[1]}")
            else:
                messagebox.showwarning("Bulunamadı", f"'{barkod}' barkodlu kitap bulunamadı.", parent=top)
        
        entry.bind("<Return>", ara)
        tk.Button(top, text="ARA", command=ara, font=("Segoe UI", 11, "bold"),
                  bg="#27AE60", fg="white", width=15).pack(pady=15)
    
    # --- PAROLA DEĞİŞTİRME ---
    def parola_degistir_penceresi(self):
        """Öğretmen ve öğrenci parolalarını değiştirme penceresi"""
        top = tk.Toplevel(self.root)
        top.title("Parola Değiştir")
        top.geometry("380x320")
        top.configure(bg="#1a1a2e")
        top.resizable(False, False)
        
        tk.Label(top, text="🔑 Parola Yönetimi", font=("Arial", 14, "bold"),
                 bg="#1a1a2e", fg="#e94560").pack(pady=15)
        
        # Öğretmen parolası
        frame1 = tk.Frame(top, bg="#1a1a2e")
        frame1.pack(fill=tk.X, padx=30, pady=10)
        tk.Label(frame1, text="Öğretmen Parolası:", bg="#1a1a2e", fg="white",
                 font=("Arial", 10), width=18, anchor="w").pack(side=tk.LEFT)
        ogretmen_entry = tk.Entry(frame1, font=("Arial", 11), width=18)
        ogretmen_entry.pack(side=tk.RIGHT)
        
        # Mevcut parolaları yükle
        self.cursor.execute("SELECT parola FROM parolalar WHERE tip='ogretmen'")
        result = self.cursor.fetchone()
        if result:
            ogretmen_entry.insert(0, result[0])
        
        # Öğrenci parolası
        frame2 = tk.Frame(top, bg="#1a1a2e")
        frame2.pack(fill=tk.X, padx=30, pady=10)
        tk.Label(frame2, text="Öğrenci Parolası:", bg="#1a1a2e", fg="white",
                 font=("Arial", 10), width=18, anchor="w").pack(side=tk.LEFT)
        ogrenci_entry = tk.Entry(frame2, font=("Arial", 11), width=18)
        ogrenci_entry.pack(side=tk.RIGHT)
        
        self.cursor.execute("SELECT parola FROM parolalar WHERE tip='ogrenci'")
        result = self.cursor.fetchone()
        if result:
            ogrenci_entry.insert(0, result[0])
        
        def kaydet():
            yeni_ogretmen = ogretmen_entry.get().strip()
            yeni_ogrenci = ogrenci_entry.get().strip()
            
            if not yeni_ogretmen or not yeni_ogrenci:
                messagebox.showerror("Hata", "Parolalar boş bırakılamaz!", parent=top)
                return
            
            if len(yeni_ogretmen) < 4 or len(yeni_ogrenci) < 4:
                messagebox.showerror("Hata", "Parolalar en az 4 karakter olmalı!", parent=top)
                return
            
            self.cursor.execute("UPDATE parolalar SET parola=? WHERE tip='ogretmen'", (yeni_ogretmen,))
            self.cursor.execute("UPDATE parolalar SET parola=? WHERE tip='ogrenci'", (yeni_ogrenci,))
            self.conn.commit()
            
            messagebox.showinfo("Başarılı", "Parolalar güncellendi!", parent=top)
            top.destroy()
        
        tk.Button(top, text="💾 KAYDET", command=kaydet, font=("Arial", 11, "bold"),
                  bg="#e94560", fg="white", width=15, bd=0, cursor="hand2").pack(pady=25)
        
        tk.Label(top, text="Not: Değişiklikler bir sonraki girişte\ngeçerli olacaktır.", 
                 bg="#1a1a2e", fg="#666", font=("Arial", 9)).pack()
    
    # --- ÖĞRENCİ EXCEL IMPORT ---
    def ogrenci_excel_import(self):
        """Excel dosyasından öğrenci aktarma"""
        if not EXCEL_DESTEGI:
            messagebox.showerror("Hata", "Excel desteği için 'openpyxl' yükleyin")
            return
        
        dosya = filedialog.askopenfilename(
            title="Öğrenci Excel Dosyası Seç",
            filetypes=[("Excel Dosyaları", "*.xlsx *.xls"), ("Tüm Dosyalar", "*.*")]
        )
        
        if not dosya:
            return
        
        try:
            wb = load_workbook(dosya)
            ws = wb.active
            
            # Başlıkları bul
            basliklar = []
            for cell in ws[1]:
                basliklar.append(str(cell.value).lower().strip() if cell.value else "")
            
            # Sütun eşleştirme
            sutun_map = {'okul_no': None, 'ad_soyad': None, 'sinif': None}
            
            for i, baslik in enumerate(basliklar):
                if 'no' in baslik or 'numara' in baslik:
                    sutun_map['okul_no'] = i
                elif 'ad' in baslik or 'isim' in baslik or 'soyad' in baslik:
                    if sutun_map['ad_soyad'] is None:
                        sutun_map['ad_soyad'] = i
                elif 'sınıf' in baslik or 'sinif' in baslik:
                    sutun_map['sinif'] = i
            
            # Verileri oku
            ogrenciler = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                
                ogrenci = {
                    'okul_no': str(row[sutun_map['okul_no']]).strip() if sutun_map['okul_no'] is not None and row[sutun_map['okul_no']] else "",
                    'ad_soyad': str(row[sutun_map['ad_soyad']]).strip() if sutun_map['ad_soyad'] is not None and row[sutun_map['ad_soyad']] else "",
                    'sinif': str(row[sutun_map['sinif']]).strip() if sutun_map['sinif'] is not None and row[sutun_map['sinif']] else ""
                }
                
                if ogrenci['okul_no'] and ogrenci['ad_soyad']:
                    ogrenciler.append(ogrenci)
            
            if not ogrenciler:
                messagebox.showwarning("Uyarı", "Excel dosyasında geçerli öğrenci bulunamadı.")
                return
            
            onay = messagebox.askyesno("Onay", 
                f"{len(ogrenciler)} öğrenci bulundu.\n\nBu öğrencileri veritabanına eklemek istiyor musunuz?")
            
            if onay:
                eklenen = 0
                guncellenen = 0
                for ogr in ogrenciler:
                    try:
                        self.cursor.execute(
                            "INSERT OR REPLACE INTO ogrenciler (okul_no, ad_soyad, sinif) VALUES (?,?,?)",
                            (ogr['okul_no'], ogr['ad_soyad'], ogr['sinif'])
                        )
                        eklenen += 1
                    except Exception as e:
                        print(f"Öğrenci eklenemedi: {ogr['ad_soyad']} - {e}")
                
                self.conn.commit()
                messagebox.showinfo("Başarılı", f"{eklenen} öğrenci aktarıldı.")
                self.durum_yaz(f"Excel'den {eklenen} öğrenci aktarıldı.")
                
        except Exception as e:
            messagebox.showerror("Hata", f"Excel dosyası okunamadı:\n{e}")
    
    # --- ÖĞRENCİ LİSTESİ ---
    def ogrenci_listesi_penceresi(self):
        """Öğrenci listesi ve kitap takip penceresi"""
        top = tk.Toplevel(self.root)
        top.title("Öğrenci Listesi")
        top.geometry("900x600")
        top.configure(bg="#1a1a2e")
        
        # Başlık
        header = tk.Frame(top, bg="#1a1a2e")
        header.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(header, text="👥 Öğrenci Listesi", font=("Arial", 16, "bold"),
                 bg="#1a1a2e", fg="#e94560").pack(side=tk.LEFT)
        
        # Arama
        search_frame = tk.Frame(header, bg="#1a1a2e")
        search_frame.pack(side=tk.RIGHT)
        
        tk.Label(search_frame, text="🔍 Ara:", bg="#1a1a2e", fg="white",
                 font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
        
        search_entry = tk.Entry(search_frame, font=("Arial", 11), width=25)
        search_entry.pack(side=tk.LEFT, padx=5)
        
        # Tablo çerçevesi
        table_frame = tk.Frame(top, bg="#16213e")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Treeview
        columns = ("OkulNo", "AdSoyad", "Sinif", "AldigiKitap")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        
        tree.heading("OkulNo", text="Okul No")
        tree.heading("AdSoyad", text="Ad Soyad")
        tree.heading("Sinif", text="Sınıf")
        tree.heading("AldigiKitap", text="Aldığı Kitap")
        
        tree.column("OkulNo", width=100, anchor="center")
        tree.column("AdSoyad", width=200)
        tree.column("Sinif", width=80, anchor="center")
        tree.column("AldigiKitap", width=300)
        
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Satır renkleri
        tree.tag_configure('kitapli', background='#FDEBD0')
        tree.tag_configure('normal', background='#ffffff')
        
        def listeyi_yukle(arama=""):
            tree.delete(*tree.get_children())
            
            if arama:
                query = """
                    SELECT o.okul_no, o.ad_soyad, o.sinif, k.ad 
                    FROM ogrenciler o 
                    LEFT JOIN odunc_alanlar oa ON o.okul_no = oa.ogrenci_no 
                    LEFT JOIN kitaplar k ON oa.kitap_id = k.id
                    WHERE o.okul_no LIKE ? OR o.ad_soyad LIKE ? OR o.sinif LIKE ?
                    ORDER BY o.sinif, o.ad_soyad
                """
                self.cursor.execute(query, (f"%{arama}%", f"%{arama}%", f"%{arama}%"))
            else:
                query = """
                    SELECT o.okul_no, o.ad_soyad, o.sinif, k.ad 
                    FROM ogrenciler o 
                    LEFT JOIN odunc_alanlar oa ON o.okul_no = oa.ogrenci_no 
                    LEFT JOIN kitaplar k ON oa.kitap_id = k.id
                    ORDER BY o.sinif, o.ad_soyad
                """
                self.cursor.execute(query)
            
            for row in self.cursor.fetchall():
                kitap = row[3] if row[3] else "-"
                tag = 'kitapli' if row[3] else 'normal'
                tree.insert("", tk.END, values=(row[0], row[1], row[2], kitap), tags=(tag,))
        
        def arama_yap(event=None):
            listeyi_yukle(search_entry.get())
        
        search_entry.bind("<KeyRelease>", arama_yap)
        
        # Çift tık ile öğrenci geçmişi
        def gecmis_goster(event):
            secili = tree.selection()
            if secili:
                okul_no = tree.item(secili)['values'][0]
                self.ogrenci_gecmisi(okul_no)
        
        tree.bind("<Double-1>", gecmis_goster)
        
        # İstatistik
        stat_frame = tk.Frame(top, bg="#1a1a2e")
        stat_frame.pack(fill=tk.X, padx=20, pady=10)
        
        self.cursor.execute("SELECT COUNT(*) FROM ogrenciler")
        toplam = self.cursor.fetchone()[0]
        
        self.cursor.execute("SELECT COUNT(DISTINCT ogrenci_no) FROM odunc_alanlar")
        kitapli = self.cursor.fetchone()[0]
        
        tk.Label(stat_frame, text=f"Toplam: {toplam} öğrenci | Kitap alanlar: {kitapli}",
                 bg="#1a1a2e", fg="#aaa", font=("Arial", 10)).pack()
        
        # İlk yükleme
        listeyi_yukle()
    
    # ========================================
    # TOPLU İŞLEMLER VE YORUMLAR
    # ========================================
    
    def toplu_iade(self, secili_items):
        """Birden fazla kitabı toplu iade al"""
        if not secili_items:
            messagebox.showwarning("Seçim Yok", "Lütfen iade alınacak kitapları seçin.")
            return
        
        # Sadece ödünçte olan kitapları filtrele
        iade_edilecek = []
        for item in secili_items:
            values = self.tree.item(item)['values']
            if values[6] != 'Mevcut':  # Durum
                iade_edilecek.append((item, values[0], values[1]))  # (item, id, ad)
        
        if not iade_edilecek:
            messagebox.showinfo("Bilgi", "Seçili kitaplar zaten mevcut durumda.")
            return
        
        onay = messagebox.askyesno("Toplu İade", 
            f"{len(iade_edilecek)} kitap iade alınacak.\n\nDevam etmek istiyor musunuz?\n\n(Not: Toplu iadede yorum istenmez)")
        
        if not onay:
            return
        
        iade_sayisi = 0
        for item, kitap_id, kitap_ad in iade_edilecek:
            try:
                # Verileri al
                self.cursor.execute("""
                    SELECT k.ad, k.yazar, k.tur, o.ogrenci_ad, o.ogrenci_no, o.sinif, o.alinma_tarihi
                    FROM odunc_alanlar o 
                    JOIN kitaplar k ON o.kitap_id = k.id 
                    WHERE o.kitap_id = ?""", (kitap_id,))
                veri = self.cursor.fetchone()
                
                if veri:
                    bugun = datetime.now().strftime("%d.%m.%Y")
                    self.cursor.execute("""
                        INSERT INTO odunc_gecmisi (kitap_ad, yazar, tur, ogrenci_ad, ogrenci_no, sinif, alinma_tarihi, iade_tarihi)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (veri[0], veri[1], veri[2], veri[3], veri[4], veri[5], veri[6], bugun))
                
                # İade işlemi
                self.cursor.execute("UPDATE kitaplar SET durum='Mevcut' WHERE id=?", (kitap_id,))
                self.cursor.execute("DELETE FROM odunc_alanlar WHERE kitap_id=?", (kitap_id,))
                iade_sayisi += 1
            except Exception as e:
                print(f"İade hatası ({kitap_ad}): {e}")
        
        self.conn.commit()
        self.verileri_guncelle()
        self.durum_yaz(f"{iade_sayisi} kitap iade alındı.")
        messagebox.showinfo("Başarılı", f"{iade_sayisi} kitap başarıyla iade alındı.")
    
    def toplu_odunc(self):
        """Birden fazla kitabı aynı öğrenciye ödünç ver"""
        secili = self.tree.selection()
        if not secili or len(secili) < 2:
            messagebox.showwarning("Seçim Yok", "Lütfen en az 2 kitap seçin.\n\nİpucu: Ctrl tuşuna basılı tutarak çoklu seçim yapabilirsiniz.")
            return
        
        # Sadece mevcut kitapları filtrele
        odunc_verilecek = []
        for item in secili:
            values = self.tree.item(item)['values']
            if values[6] == 'Mevcut':
                odunc_verilecek.append((values[0], values[1]))  # (id, ad)
        
        if not odunc_verilecek:
            messagebox.showinfo("Bilgi", "Seçili kitapların tamamı zaten ödünçte.")
            return
        
        # Öğrenci bilgileri penceresi
        top = tk.Toplevel(self.root)
        top.title("📚 Toplu Ödünç Verme")
        top.geometry("450x400")
        top.configure(bg=self.bg_color)
        
        tk.Label(top, text="📚 Toplu Ödünç Verme", font=("Segoe UI", 14, "bold"),
                 bg=self.bg_color, fg=self.panel_color).pack(pady=15)
        
        tk.Label(top, text=f"{len(odunc_verilecek)} kitap seçildi", font=("Segoe UI", 11),
                 bg=self.bg_color, fg="#666").pack()
        
        # Kitap listesi
        kitap_list = tk.Listbox(top, height=5, font=("Segoe UI", 9))
        kitap_list.pack(fill=tk.X, padx=30, pady=10)
        for kitap_id, kitap_ad in odunc_verilecek:
            kitap_list.insert(tk.END, f"• {kitap_ad[:40]}")
        
        tk.Label(top, text="Öğrenci Bilgileri:", font=("Segoe UI", 11, "bold"),
                 bg=self.bg_color).pack(pady=(10, 5))
        
        entries = {}
        for alan in ["Öğrenci Adı Soyadı", "Okul No", "Sınıf"]:
            frame = tk.Frame(top, bg=self.bg_color)
            frame.pack(fill=tk.X, padx=30, pady=3)
            tk.Label(frame, text=alan, bg=self.bg_color, width=18, anchor="w").pack(side=tk.LEFT)
            e = ttk.Entry(frame)
            e.pack(side=tk.RIGHT, fill=tk.X, expand=True)
            entries[alan] = e
        
        def onayla():
            if not all(e.get() for e in entries.values()):
                messagebox.showwarning("Eksik", "Tüm alanları doldurunuz.", parent=top)
                return
            
            # Süre hesapla
            self.cursor.execute("SELECT deger FROM ayarlar WHERE anahtar='odunc_suresi'")
            gun = int(self.cursor.fetchone()[0])
            tarih_fmt = "%d.%m.%Y"
            bugun = datetime.now().strftime(tarih_fmt)
            iade = (datetime.now() + timedelta(days=gun)).strftime(tarih_fmt)
            
            ogrenci_ad = entries["Öğrenci Adı Soyadı"].get()
            ogrenci_no = entries["Okul No"].get()
            sinif = entries["Sınıf"].get()
            
            verilen = 0
            for kitap_id, kitap_ad in odunc_verilecek:
                try:
                    self.cursor.execute("UPDATE kitaplar SET durum='Ödünç Verildi' WHERE id=?", (kitap_id,))
                    self.cursor.execute("""
                        INSERT INTO odunc_alanlar (kitap_id, ogrenci_ad, ogrenci_no, sinif, alinma_tarihi, iade_tarihi)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (kitap_id, ogrenci_ad, ogrenci_no, sinif, bugun, iade))
                    verilen += 1
                except Exception as e:
                    print(f"Ödünç verme hatası: {e}")
            
            self.conn.commit()
            top.destroy()
            self.verileri_guncelle()
            self.durum_yaz(f"{verilen} kitap {ogrenci_ad}'a verildi.")
            messagebox.showinfo("Başarılı", f"{verilen} kitap başarıyla ödünç verildi.\n\nSon iade tarihi: {iade}")
        
        tk.Button(top, text="✅ ÖDÜNÇ VER", command=onayla,
                  font=("Segoe UI", 11, "bold"), bg=self.action_color, fg="white",
                  width=15, cursor="hand2").pack(pady=20)
    
    def yorumlari_goster(self):
        """Seçili kitabın yorumlarını göster"""
        secili = self.tree.selection()
        if not secili:
            messagebox.showwarning("Seçim Yok", "Lütfen bir kitap seçin.")
            return
        
        item = self.tree.item(secili[0])
        kitap_id = item['values'][0]
        kitap_ad = item['values'][1]
        
        # Yorumları getir
        self.cursor.execute("""
            SELECT ogrenci_ad, puan, yorum, tarih 
            FROM yorumlar 
            WHERE kitap_id = ? 
            ORDER BY id DESC
        """, (kitap_id,))
        yorumlar = self.cursor.fetchall()
        
        # Pencere
        top = tk.Toplevel(self.root)
        top.title(f"💬 Yorumlar - {kitap_ad[:30]}")
        top.geometry("500x450")
        top.configure(bg="#1a1a2e")
        
        tk.Label(top, text="💬 Kitap Yorumları", font=("Segoe UI", 14, "bold"),
                 bg="#1a1a2e", fg="#3b82f6").pack(pady=15)
        
        tk.Label(top, text=f"'{kitap_ad[:40]}'", font=("Segoe UI", 11),
                 bg="#1a1a2e", fg="white", wraplength=450).pack()
        
        # Ortalama puan
        if yorumlar:
            ort_puan = sum(y[1] for y in yorumlar) / len(yorumlar)
            yildiz = "⭐" * round(ort_puan)
            tk.Label(top, text=f"{yildiz} ({ort_puan:.1f}/5) - {len(yorumlar)} değerlendirme",
                     bg="#1a1a2e", fg="#f59e0b", font=("Segoe UI", 11)).pack(pady=10)
        
        # Yorumlar listesi
        frame = tk.Frame(top, bg="#16213e")
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        canvas = tk.Canvas(frame, bg="#16213e", highlightthickness=0)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#16213e")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        if yorumlar:
            for ogrenci, puan, yorum, tarih in yorumlar:
                yorum_frame = tk.Frame(scrollable_frame, bg="#0f3460", pady=10, padx=10)
                yorum_frame.pack(fill=tk.X, pady=5, padx=5)
                
                header = tk.Frame(yorum_frame, bg="#0f3460")
                header.pack(fill=tk.X)
                
                tk.Label(header, text=f"👤 {ogrenci}", font=("Segoe UI", 10, "bold"),
                         bg="#0f3460", fg="white").pack(side=tk.LEFT)
                tk.Label(header, text="⭐" * puan, font=("Segoe UI", 9),
                         bg="#0f3460", fg="#f59e0b").pack(side=tk.LEFT, padx=10)
                tk.Label(header, text=tarih, font=("Segoe UI", 9),
                         bg="#0f3460", fg="#888").pack(side=tk.RIGHT)
                
                tk.Label(yorum_frame, text=yorum, font=("Segoe UI", 10),
                         bg="#0f3460", fg="#ddd", wraplength=420, justify="left").pack(fill=tk.X, pady=(5, 0))
        else:
            tk.Label(scrollable_frame, text="Henüz yorum yapılmamış.\n\nKitabı iade eden ilk öğrenci yorum bırakabilir!",
                     bg="#16213e", fg="#888", font=("Segoe UI", 11)).pack(pady=50)
        
        tk.Button(top, text="Kapat", command=top.destroy,
                  font=("Segoe UI", 10), bg="#95A5A6", fg="white", width=10).pack(pady=10)
    
    # ========================================
    # YENİ PROFESYONEL ÖZELLİKLER
    # ========================================
    
    # --- KİTAP DÜZENLEME ---
    def kitap_duzenle_penceresi(self):
        """Seçili kitabı düzenleme penceresi"""
        secili = self.tree.selection()
        if not secili:
            messagebox.showwarning("Seçim Yok", "Lütfen düzenlemek istediğiniz kitabı seçin.")
            return
        
        item = self.tree.item(secili)
        kitap_id = item['values'][0]
        
        # Mevcut verileri al
        self.cursor.execute("SELECT ad, yazar, tur, sayfa_sayisi, raf_no, adet FROM kitaplar WHERE id=?", (kitap_id,))
        kitap = self.cursor.fetchone()
        
        if not kitap:
            messagebox.showerror("Hata", "Kitap bulunamadı.")
            return
        
        top = tk.Toplevel(self.root)
        top.title(f"Kitap Düzenle - {kitap[0][:30]}")
        top.geometry("450x450")
        top.configure(bg=self.bg_color)
        top.resizable(False, False)
        
        tk.Label(top, text="✏️ Kitap Düzenle", font=("Segoe UI", 16, "bold"),
                 bg=self.bg_color, fg=self.panel_color).pack(pady=20)
        
        form_frame = tk.Frame(top, bg=self.bg_color)
        form_frame.pack(fill=tk.BOTH, expand=True, padx=30)
        
        entries = {}
        fields = [
            ("Kitap Adı", kitap[0]),
            ("Yazar", kitap[1]),
            ("Tür", kitap[2] or ""),
            ("Sayfa Sayısı", kitap[3] or ""),
            ("Raf No", kitap[4] or ""),
            ("Adet", kitap[5] or 1)
        ]
        
        for label_text, default_value in fields:
            frame = tk.Frame(form_frame, bg=self.bg_color)
            frame.pack(fill=tk.X, pady=5)
            tk.Label(frame, text=label_text, bg=self.bg_color, width=12, anchor="w",
                     font=("Segoe UI", 10)).pack(side=tk.LEFT)
            entry = ttk.Entry(frame, font=("Segoe UI", 11))
            entry.pack(side=tk.RIGHT, fill=tk.X, expand=True)
            entry.insert(0, str(default_value))
            entries[label_text] = entry
        
        def kaydet():
            yeni_ad = entries["Kitap Adı"].get().strip()
            yeni_yazar = entries["Yazar"].get().strip()
            yeni_tur = entries["Tür"].get().strip()
            yeni_sayfa = entries["Sayfa Sayısı"].get().strip()
            yeni_raf = entries["Raf No"].get().strip()
            yeni_adet = entries["Adet"].get().strip()
            
            if not yeni_ad or not yeni_yazar:
                messagebox.showwarning("Eksik Bilgi", "Kitap Adı ve Yazar zorunludur.", parent=top)
                return
            
            try:
                sayfa_int = int(yeni_sayfa) if yeni_sayfa else None
                adet_int = int(yeni_adet) if yeni_adet else 1
            except ValueError:
                messagebox.showerror("Hata", "Sayfa sayısı ve adet sayı olmalıdır.", parent=top)
                return
            
            self.cursor.execute("""
                UPDATE kitaplar 
                SET ad=?, yazar=?, tur=?, sayfa_sayisi=?, raf_no=?, adet=?
                WHERE id=?
            """, (yeni_ad, yeni_yazar, yeni_tur, sayfa_int, yeni_raf, adet_int, kitap_id))
            self.conn.commit()
            
            top.destroy()
            self.verileri_guncelle()
            self.durum_yaz(f"'{yeni_ad}' kitabı güncellendi.")
            messagebox.showinfo("Başarılı", "Kitap bilgileri güncellendi.")
        
        btn_frame = tk.Frame(top, bg=self.bg_color)
        btn_frame.pack(fill=tk.X, padx=30, pady=20)
        
        tk.Button(btn_frame, text="💾 KAYDET", command=kaydet, font=("Segoe UI", 11, "bold"),
                  bg=self.action_color, fg="white", width=12, cursor="hand2").pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="❌ İPTAL", command=top.destroy, font=("Segoe UI", 11),
                  bg="#95A5A6", fg="white", width=12, cursor="hand2").pack(side=tk.RIGHT, padx=5)
    
    # --- İSTATİSTİK DASHBOARD ---
    def istatistik_dashboard(self):
        """Gelişmiş istatistik dashboard penceresi"""
        top = tk.Toplevel(self.root)
        top.title("📊 İstatistik Dashboard")
        top.geometry("1000x700")
        top.configure(bg="#1a1a2e")
        
        # Başlık
        tk.Label(top, text="📊 KÜTÜPHANE İSTATİSTİKLERİ", font=("Segoe UI", 20, "bold"),
                 bg="#1a1a2e", fg="#3b82f6").pack(pady=20)
        
        # Ana içerik çerçevesi
        content = tk.Frame(top, bg="#1a1a2e")
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Sol: Sayısal istatistikler
        left_frame = tk.Frame(content, bg="#16213e", width=300)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10)
        left_frame.pack_propagate(False)
        
        tk.Label(left_frame, text="📈 ÖZET", font=("Segoe UI", 14, "bold"),
                 bg="#16213e", fg="#e94560").pack(pady=15)
        
        # İstatistikleri hesapla
        self.cursor.execute("SELECT COUNT(*) FROM kitaplar")
        toplam_kitap = self.cursor.fetchone()[0]
        
        self.cursor.execute("SELECT COUNT(*) FROM kitaplar WHERE durum != 'Mevcut'")
        oduncte = self.cursor.fetchone()[0]
        
        self.cursor.execute("SELECT COUNT(*) FROM ogrenciler")
        toplam_ogrenci = self.cursor.fetchone()[0]
        
        self.cursor.execute("SELECT COUNT(*) FROM odunc_gecmisi")
        toplam_islem = self.cursor.fetchone()[0]
        
        # Bu ay
        bu_ay = datetime.now().strftime("%m.%Y")
        self.cursor.execute("SELECT COUNT(*) FROM odunc_gecmisi WHERE alinma_tarihi LIKE ?", (f"%.{bu_ay}",))
        bu_ay_islem = self.cursor.fetchone()[0]
        
        # Gecikmiş
        bugun = datetime.now()
        self.cursor.execute("SELECT COUNT(*) FROM odunc_alanlar WHERE iade_tarihi IS NOT NULL")
        self.cursor.execute("SELECT iade_tarihi FROM odunc_alanlar")
        gecikmis = 0
        for row in self.cursor.fetchall():
            if row[0]:
                try:
                    iade = datetime.strptime(row[0], "%d.%m.%Y")
                    if iade < bugun:
                        gecikmis += 1
                except:
                    pass
        
        stats = [
            ("📚 Toplam Kitap", toplam_kitap, "#3b82f6"),
            ("📖 Ödünçte", oduncte, "#f59e0b"),
            ("👥 Toplam Öğrenci", toplam_ogrenci, "#10b981"),
            ("📋 Toplam İşlem", toplam_islem, "#8b5cf6"),
            ("📅 Bu Ay", bu_ay_islem, "#06b6d4"),
            ("⚠️ Gecikmiş", gecikmis, "#ef4444"),
        ]
        
        for label, value, color in stats:
            card = tk.Frame(left_frame, bg="#0f3460", pady=10)
            card.pack(fill=tk.X, padx=10, pady=5)
            tk.Label(card, text=label, bg="#0f3460", fg="#aaa", font=("Segoe UI", 10)).pack()
            tk.Label(card, text=str(value), bg="#0f3460", fg=color, font=("Segoe UI", 24, "bold")).pack()
        
        # Sağ: Grafikler (matplotlib varsa)
        right_frame = tk.Frame(content, bg="#16213e")
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10)
        
        if MATPLOTLIB_DESTEGI:
            # Tür dağılımı grafiği
            self.cursor.execute("SELECT tur, COUNT(*) FROM kitaplar WHERE tur IS NOT NULL AND tur != '' GROUP BY tur ORDER BY COUNT(*) DESC LIMIT 6")
            tur_data = self.cursor.fetchall()
            
            if tur_data:
                fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(8, 4), facecolor='#16213e')
                
                # Pasta grafik - Tür dağılımı
                labels = [row[0][:15] for row in tur_data]
                sizes = [row[1] for row in tur_data]
                colors = ['#3b82f6', '#ef4444', '#10b981', '#f59e0b', '#8b5cf6', '#06b6d4']
                
                ax1.pie(sizes, labels=labels, colors=colors[:len(sizes)], autopct='%1.0f%%',
                        textprops={'color': 'white', 'fontsize': 8})
                ax1.set_title('Tür Dağılımı', color='white', fontsize=12)
                ax1.set_facecolor('#16213e')
                
                # Bar grafik - Aylık ödünç
                aylar = []
                sayilar = []
                for i in range(5, -1, -1):
                    tarih = datetime.now() - timedelta(days=i*30)
                    ay_str = tarih.strftime("%m.%Y")
                    self.cursor.execute("SELECT COUNT(*) FROM odunc_gecmisi WHERE alinma_tarihi LIKE ?", (f"%.{ay_str}",))
                    sayilar.append(self.cursor.fetchone()[0])
                    aylar.append(tarih.strftime("%b"))
                
                ax2.bar(aylar, sayilar, color='#3b82f6')
                ax2.set_title('Aylık Ödünç', color='white', fontsize=12)
                ax2.set_facecolor('#16213e')
                ax2.tick_params(colors='white')
                for spine in ax2.spines.values():
                    spine.set_color('#16213e')
                
                fig.tight_layout()
                
                canvas = FigureCanvasTkAgg(fig, right_frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, pady=10)
        else:
            tk.Label(right_frame, text="📊 Grafikler için matplotlib yükleyin:\npip install matplotlib",
                     bg="#16213e", fg="#aaa", font=("Segoe UI", 12)).pack(pady=50)
        
        # Alt: En popüler kitaplar
        bottom_frame = tk.Frame(top, bg="#16213e")
        bottom_frame.pack(fill=tk.X, padx=20, pady=10)
        
        tk.Label(bottom_frame, text="🔥 EN POPÜLER KİTAPLAR", font=("Segoe UI", 12, "bold"),
                 bg="#16213e", fg="#e94560").pack(pady=10)
        
        self.cursor.execute("""
            SELECT kitap_ad, COUNT(*) as sayi FROM odunc_gecmisi 
            GROUP BY kitap_ad ORDER BY sayi DESC LIMIT 5
        """)
        populer = self.cursor.fetchall()
        
        if populer:
            for i, (kitap, sayi) in enumerate(populer, 1):
                tk.Label(bottom_frame, text=f"{i}. {kitap[:40]} - {sayi} kez okundu",
                         bg="#16213e", fg="white", font=("Segoe UI", 10)).pack(anchor="w", padx=20)
        else:
            tk.Label(bottom_frame, text="Henüz yeterli veri yok.", bg="#16213e", fg="#aaa").pack()
    
    # --- SINIF SIRALAMASI ---
    def sinif_siralamasi(self):
        """Sınıfların okuma sıralaması"""
        top = tk.Toplevel(self.root)
        top.title("🏆 Sınıf Sıralaması")
        top.geometry("500x500")
        top.configure(bg="#1a1a2e")
        
        tk.Label(top, text="🏆 SINIF OKUMA YARIŞMASI", font=("Segoe UI", 18, "bold"),
                 bg="#1a1a2e", fg="#f59e0b").pack(pady=20)
        
        # Sınıflara göre okunan kitap sayısı
        self.cursor.execute("""
            SELECT sinif, COUNT(*) as toplam 
            FROM odunc_gecmisi 
            WHERE sinif IS NOT NULL AND sinif != ''
            GROUP BY sinif 
            ORDER BY toplam DESC
        """)
        siniflar = self.cursor.fetchall()
        
        if siniflar:
            for i, (sinif, toplam) in enumerate(siniflar, 1):
                renk = "#FFD700" if i == 1 else "#C0C0C0" if i == 2 else "#CD7F32" if i == 3 else "#aaa"
                emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
                
                card = tk.Frame(top, bg="#16213e", pady=10)
                card.pack(fill=tk.X, padx=30, pady=5)
                
                tk.Label(card, text=f"{emoji} {sinif}", font=("Segoe UI", 14, "bold"),
                         bg="#16213e", fg=renk).pack(side=tk.LEFT, padx=20)
                tk.Label(card, text=f"{toplam} kitap", font=("Segoe UI", 12),
                         bg="#16213e", fg="white").pack(side=tk.RIGHT, padx=20)
        else:
            tk.Label(top, text="Henüz yeterli veri yok.", bg="#1a1a2e", fg="#aaa",
                     font=("Segoe UI", 12)).pack(pady=50)
    
    # --- LİDERLİK TABLOSU ---
    def liderlik_tablosu(self):
        """Kitap kurdu liderlik tablosu"""
        top = tk.Toplevel(self.root)
        top.title("👑 Kitap Kurdu Liderlik Tablosu")
        top.geometry("600x600")
        top.configure(bg="#1a1a2e")
        
        tk.Label(top, text="👑 KİTAP KURDU LİDERLİK TABLOSU", font=("Segoe UI", 18, "bold"),
                 bg="#1a1a2e", fg="#f59e0b").pack(pady=20)
        
        # En çok kitap okuyan öğrenciler
        self.cursor.execute("""
            SELECT ogrenci_ad, sinif, COUNT(*) as toplam, SUM(
                CASE 
                    WHEN julianday(iade_tarihi) - julianday(alinma_tarihi) <= 7 THEN 15
                    WHEN julianday(iade_tarihi) - julianday(alinma_tarihi) <= 14 THEN 12
                    ELSE 10
                END
            ) as puan
            FROM odunc_gecmisi 
            WHERE ogrenci_ad IS NOT NULL
            GROUP BY ogrenci_ad 
            ORDER BY toplam DESC 
            LIMIT 10
        """)
        liderler = self.cursor.fetchall()
        
        # Tablo
        tree_frame = tk.Frame(top, bg="#16213e")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        columns = ("Sira", "Ogrenci", "Sinif", "Kitap", "Puan")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=10)
        
        tree.heading("Sira", text="#")
        tree.heading("Ogrenci", text="Öğrenci")
        tree.heading("Sinif", text="Sınıf")
        tree.heading("Kitap", text="Kitap Sayısı")
        tree.heading("Puan", text="Puan")
        
        tree.column("Sira", width=50, anchor="center")
        tree.column("Ogrenci", width=200)
        tree.column("Sinif", width=80, anchor="center")
        tree.column("Kitap", width=100, anchor="center")
        tree.column("Puan", width=80, anchor="center")
        
        tree.pack(fill=tk.BOTH, expand=True)
        
        for i, (ad, sinif, toplam, puan) in enumerate(liderler, 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else str(i)
            puan_val = puan if puan else toplam * 10
            tree.insert("", tk.END, values=(emoji, ad, sinif or "-", toplam, puan_val))
    
    # --- PDF RAPOR ---
    def pdf_rapor_olustur(self):
        """PDF formatında rapor oluştur"""
        if not PDF_DESTEGI:
            messagebox.showerror("Hata", "PDF desteği için fpdf2 yükleyin:\npip install fpdf2")
            return
        
        dosya = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Dosyası", "*.pdf")],
            initialfile=f"Kutuphane_Rapor_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
        
        if not dosya:
            return
        
        try:
            pdf = FPDF()
            pdf.add_page()
            
            # Türkçe karakter desteği için font
            pdf.add_font('DejaVu', '', '/System/Library/Fonts/Supplemental/Arial Unicode.ttf', uni=True)
            pdf.set_font('DejaVu', '', 12)
            
            # Başlık
            pdf.set_font('DejaVu', '', 20)
            pdf.cell(0, 15, 'ŞAİK KÜTÜPHANE RAPORU', ln=True, align='C')
            pdf.set_font('DejaVu', '', 10)
            pdf.cell(0, 10, f'Tarih: {datetime.now().strftime("%d.%m.%Y %H:%M")}', ln=True, align='C')
            pdf.ln(10)
            
            # İstatistikler
            pdf.set_font('DejaVu', '', 14)
            pdf.cell(0, 10, 'GENEL İSTATİSTİKLER', ln=True)
            pdf.set_font('DejaVu', '', 11)
            
            self.cursor.execute("SELECT COUNT(*) FROM kitaplar")
            pdf.cell(0, 8, f'Toplam Kitap: {self.cursor.fetchone()[0]}', ln=True)
            
            self.cursor.execute("SELECT COUNT(*) FROM kitaplar WHERE durum != 'Mevcut'")
            pdf.cell(0, 8, f'Ödünçte: {self.cursor.fetchone()[0]}', ln=True)
            
            self.cursor.execute("SELECT COUNT(*) FROM ogrenciler")
            pdf.cell(0, 8, f'Kayıtlı Öğrenci: {self.cursor.fetchone()[0]}', ln=True)
            
            pdf.ln(10)
            
            # Ödünçteki kitaplar
            pdf.set_font('DejaVu', '', 14)
            pdf.cell(0, 10, 'ÖDÜNÇTEKİ KİTAPLAR', ln=True)
            pdf.set_font('DejaVu', '', 9)
            
            self.cursor.execute("""
                SELECT k.ad, o.ogrenci_ad, o.sinif, o.iade_tarihi 
                FROM odunc_alanlar o 
                JOIN kitaplar k ON o.kitap_id = k.id
            """)
            
            for row in self.cursor.fetchall():
                satir = f"• {row[0][:30]} - {row[1]} ({row[2]}) - İade: {row[3]}"
                pdf.cell(0, 6, satir, ln=True)
            
            pdf.output(dosya)
            messagebox.showinfo("Başarılı", f"PDF rapor oluşturuldu:\n{dosya}")
            self.durum_yaz("PDF rapor oluşturuldu.")
            
        except Exception as e:
            messagebox.showerror("Hata", f"PDF oluşturma hatası:\n{e}")
    
    # --- TEMA DEĞİŞTİR ---
    def tema_degistir(self):
        """Dark/Light tema değiştirme"""
        self.cursor.execute("SELECT deger FROM ayarlar WHERE anahtar='tema'")
        mevcut = self.cursor.fetchone()
        mevcut_tema = mevcut[0] if mevcut else "dark"
        
        yeni_tema = "light" if mevcut_tema == "dark" else "dark"
        
        self.cursor.execute("UPDATE ayarlar SET deger=? WHERE anahtar='tema'", (yeni_tema,))
        self.conn.commit()
        
        messagebox.showinfo("Tema Değiştirildi", 
            f"Tema '{yeni_tema}' olarak ayarlandı.\n\nDeğişikliğin tam olarak uygulanması için uygulamayı yeniden başlatın.")
    
    # --- GECİKME BİLDİRİMİ ---
    def ai_asistan_penceresi(self):
        """Groq API tabanlı AI Asistanı"""
        top = tk.Toplevel(self.root)
        top.title("🤖 ŞAİK Kütüphane Yapay Zeka Asistanı")
        top.geometry("800x600")
        top.configure(bg=self.bg_color)
        
        chat_frame = tk.Frame(top, bg=self.card_bg, padx=10, pady=10)
        chat_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        chat_history = tk.Text(chat_frame, font=("Segoe UI", 11), bg="#f8fafc", fg=self.text_primary, wrap=tk.WORD, state=tk.DISABLED)
        scrollbar = ttk.Scrollbar(chat_frame, command=chat_history.yview)
        chat_history.configure(yscrollcommand=scrollbar.set)
        chat_history.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        input_frame = tk.Frame(top, bg=self.bg_color)
        input_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        entry_mesaj = tk.Text(input_frame, height=3, font=("Segoe UI", 11), wrap=tk.WORD)
        entry_mesaj.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        # Önceden toplanmış konteksti hazırlama
        self.ai_system_prompt = ""
        
        def context_hazirla():
            try:
                self.cursor.execute("SELECT COUNT(*) FROM kitaplar")
                toplam_kitap = self.cursor.fetchone()[0]
                
                self.cursor.execute("SELECT COUNT(*) FROM odunc_alanlar")
                oduncte = self.cursor.fetchone()[0]
                
                self.cursor.execute("SELECT COUNT(*) FROM ogrenciler")
                ogrenci_sayisi = self.cursor.fetchone()[0]
                
                gecikmis = 0
                bugun = datetime.now()
                self.cursor.execute("SELECT iade_tarihi FROM odunc_alanlar")
                for (iade_tarihi_str,) in self.cursor.fetchall():
                    if iade_tarihi_str:
                        try:
                            iade_tarihi = datetime.strptime(iade_tarihi_str, "%d.%m.%Y")
                            if (iade_tarihi - bugun).days < 0:
                                gecikmis += 1
                        except: pass
                
                prompt = f"""Sen 'ŞAİK Kütüphane Yönetim Sistemi' nin resmi Yapay Zeka Asistanısın. 
Bu yazılım öğretmenler ve öğrenciler için geliştirildi. Şifre ve parolalar güvenlik amacıyla senden gizlenmiştir, şifreleri asla söylememelisin.
Sistem özellikleri: Barkod okuma ve oluşturma, Excel'den toplu içe/dışa aktarma, öğrenci bazlı oyunlaştırma (rozetler, liderlik panosu), Wrapped (yıl sonu özeti), oyunlar (Pong, Flappy Bird, Yılan), zeka egzersizleri, rezervasyon sistemi.
Şu anki veritabanı durumu: 
- Toplam Kitap Sayısı: {toplam_kitap}
- Ödünç Verilen Kitap: {oduncte}
- Gecikmiş İadeler: {gecikmis}
- Kayıtlı Öğrenci: {ogrenci_sayisi}

Kullanıcı öğretmen modunda seninle konuşuyor. Kütüphane, kitaplar, yazılım özellikleri veya genel bilgiler hakkında profesyonel, yardımsever ve Türkçe dilinde yanıtlar ver. Bilgileri hatasız aktar. Markdown kullan."""
                self.ai_system_prompt = prompt
                
                def update_ui():
                    chat_history.config(state=tk.NORMAL)
                    chat_history.insert(tk.END, "🤖 Asistan: Merhaba! ŞAİK Kütüphane Asistanı hazır. Size nasıl yardımcı olabilirim?\n\n")
                    chat_history.config(state=tk.DISABLED)
                    chat_history.see(tk.END)
                top.after(0, update_ui)
                
            except Exception as e:
                print("Context hazırlama hatası:", e)

        chat_context = []
        
        # API anahtarını api_key.txt dosyasından, yoksa çevre değişkeninden al
        api_key_dosyasi = os.path.join(os.path.dirname(os.path.abspath(__file__)), "api_key.txt")
        if os.path.exists(api_key_dosyasi):
            with open(api_key_dosyasi, "r", encoding="utf-8") as f:
                api_key = f.read().strip()
        else:
            api_key = os.environ.get("GROQ_API_KEY", "")
        
        def gonder(event=None):
            mesaj = entry_mesaj.get("1.0", tk.END).strip()
            if not mesaj: return 'break'
            
            entry_mesaj.delete("1.0", tk.END)
            
            chat_history.config(state=tk.NORMAL)
            chat_history.insert(tk.END, f"👤 Siz: {mesaj}\n\n")
            chat_history.see(tk.END)
            chat_history.config(state=tk.DISABLED)
            
            def api_istek():
                btn_gonder.config(state=tk.DISABLED, text="Bekleniyor...")
                try:
                    messages = [{"role": "system", "content": self.ai_system_prompt}]
                    messages.extend(chat_context)
                    messages.append({"role": "user", "content": mesaj})
                    
                    url = "https://api.groq.com/openai/v1/chat/completions"
                    headers = {
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json"
                    }
                    data = {
                        "model": "llama-3.3-70b-versatile",
                        "messages": messages,
                        "temperature": 0.7
                    }
                    
                    req = urllib.request.Request(url, headers=headers, data=json.dumps(data).encode('utf-8'))
                    with urllib.request.urlopen(req) as response:
                        res_data = json.loads(response.read().decode('utf-8'))
                        cevap = res_data['choices'][0]['message']['content']
                    
                    chat_context.append({"role": "user", "content": mesaj})
                    chat_context.append({"role": "assistant", "content": cevap})
                    
                    def arayuz_guncelle():
                        chat_history.config(state=tk.NORMAL)
                        chat_history.insert(tk.END, f"🤖 Asistan: {cevap}\n\n")
                        chat_history.see(tk.END)
                        chat_history.config(state=tk.DISABLED)
                        btn_gonder.config(state=tk.NORMAL, text="🚀 GÖNDER")
                    
                    top.after(0, arayuz_guncelle)
                        
                except Exception as e:
                    error_msg = str(e)
                    def hata_bas(msg=error_msg):
                        chat_history.config(state=tk.NORMAL)
                        chat_history.insert(tk.END, f"❌ Hata: {msg}\n\n")
                        chat_history.see(tk.END)
                        chat_history.config(state=tk.DISABLED)
                        btn_gonder.config(state=tk.NORMAL, text="🚀 GÖNDER")
                    top.after(0, hata_bas)

            threading.Thread(target=api_istek, daemon=True).start()
            return 'break'
            
        btn_gonder = ttk.Button(input_frame, text="🚀 GÖNDER", style="Islem.TButton", command=gonder)
        btn_gonder.pack(side=tk.RIGHT, fill=tk.Y)
        
        entry_mesaj.bind("<Return>", lambda e: 'break' if not getattr(e, 'state', 0) & 1 else None)
        entry_mesaj.bind("<Shift-Return>", lambda e: None)
        entry_mesaj.bind("<Key>", lambda e: gonder() if getattr(e, 'keysym', '') == 'Return' and not getattr(e, 'state', 0) & 1 else None)
        
        context_hazirla()

    def gecikme_bildirimi_goster(self):
        """Gecikmiş kitapları listele"""
        top = tk.Toplevel(self.root)
        top.title("⚠️ Gecikmiş Kitaplar")
        top.geometry("700x500")
        top.configure(bg="#1a1a2e")
        
        tk.Label(top, text="⚠️ GECİKMİŞ KİTAPLAR", font=("Segoe UI", 16, "bold"),
                 bg="#1a1a2e", fg="#ef4444").pack(pady=15)
        
        # Tablo
        tree_frame = tk.Frame(top, bg="#16213e")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        columns = ("Kitap", "Ogrenci", "Sinif", "OkulNo", "IadeTarihi", "GecikmeGun")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        tree.heading("Kitap", text="Kitap")
        tree.heading("Ogrenci", text="Öğrenci")
        tree.heading("Sinif", text="Sınıf")
        tree.heading("OkulNo", text="Okul No")
        tree.heading("IadeTarihi", text="İade Tarihi")
        tree.heading("GecikmeGun", text="Gecikme")
        
        tree.column("Kitap", width=200)
        tree.column("Ogrenci", width=150)
        tree.column("Sinif", width=60, anchor="center")
        tree.column("OkulNo", width=80, anchor="center")
        tree.column("IadeTarihi", width=90, anchor="center")
        tree.column("GecikmeGun", width=70, anchor="center")
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        bugun = datetime.now()
        self.cursor.execute("""
            SELECT k.ad, o.ogrenci_ad, o.sinif, o.ogrenci_no, o.iade_tarihi 
            FROM odunc_alanlar o 
            JOIN kitaplar k ON o.kitap_id = k.id
        """)
        
        gecikmis_sayisi = 0
        for row in self.cursor.fetchall():
            if row[4]:
                try:
                    iade = datetime.strptime(row[4], "%d.%m.%Y")
                    gecikme = (bugun - iade).days
                    if gecikme > 0:
                        tree.insert("", tk.END, values=(row[0], row[1], row[2], row[3], row[4], f"{gecikme} gün"))
                        gecikmis_sayisi += 1
                except:
                    pass
        
        # Alt bilgi
        tk.Label(top, text=f"Toplam {gecikmis_sayisi} gecikmiş kitap bulundu.",
                 bg="#1a1a2e", fg="#aaa", font=("Segoe UI", 10)).pack(pady=10)
    
    # --- REZERVASYON LİSTESİ ---
    def rezervasyon_listesi(self):
        """Aktif rezervasyonları listele"""
        top = tk.Toplevel(self.root)
        top.title("📅 Rezervasyonlar")
        top.geometry("700x450")
        top.configure(bg="#1a1a2e")
        
        tk.Label(top, text="📅 AKTİF REZERVASYONLAR", font=("Segoe UI", 16, "bold"),
                 bg="#1a1a2e", fg="#3b82f6").pack(pady=15)
        
        # Tablo
        tree_frame = tk.Frame(top, bg="#16213e")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        columns = ("ID", "Kitap", "Ogrenci", "Tarih", "Durum")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        tree.heading("ID", text="ID")
        tree.heading("Kitap", text="Kitap")
        tree.heading("Ogrenci", text="Öğrenci")
        tree.heading("Tarih", text="Tarih")
        tree.heading("Durum", text="Durum")
        
        tree.column("ID", width=50, anchor="center")
        tree.column("Kitap", width=250)
        tree.column("Ogrenci", width=150)
        tree.column("Tarih", width=100, anchor="center")
        tree.column("Durum", width=100, anchor="center")
        
        tree.pack(fill=tk.BOTH, expand=True)
        
        self.cursor.execute("""
            SELECT r.id, k.ad, r.ogrenci_ad, r.tarih, r.durum 
            FROM rezervasyonlar r 
            JOIN kitaplar k ON r.kitap_id = k.id
            WHERE r.durum = 'Bekliyor'
            ORDER BY r.tarih
        """)
        
        for row in self.cursor.fetchall():
            tree.insert("", tk.END, values=row)
        
        # İptal butonu
        def iptal_et():
            secili = tree.selection()
            if secili:
                rez_id = tree.item(secili)['values'][0]
                if messagebox.askyesno("Onay", "Bu rezervasyonu iptal etmek istiyor musunuz?"):
                    self.cursor.execute("UPDATE rezervasyonlar SET durum='İptal' WHERE id=?", (rez_id,))
                    self.conn.commit()
                    tree.delete(secili)
                    self.durum_yaz("Rezervasyon iptal edildi.")
        
        if self.kullanici_tipi == "ogretmen":
            tk.Button(top, text="❌ Seçili Rezervasyonu İptal Et", command=iptal_et,
                      font=("Segoe UI", 10), bg="#ef4444", fg="white").pack(pady=10)
    
    # --- YENİ REZERVASYON ---
    def yeni_rezervasyon(self):
        """Yeni rezervasyon oluştur"""
        top = tk.Toplevel(self.root)
        top.title("➕ Yeni Rezervasyon")
        top.geometry("400x350")
        top.configure(bg=self.bg_color)
        
        tk.Label(top, text="📅 Yeni Rezervasyon", font=("Segoe UI", 14, "bold"),
                 bg=self.bg_color, fg=self.panel_color).pack(pady=20)
        
        # Kitap seçimi
        tk.Label(top, text="Kitap:", bg=self.bg_color).pack(anchor="w", padx=30)
        self.cursor.execute("SELECT id, ad FROM kitaplar WHERE durum != 'Mevcut' ORDER BY ad")
        kitaplar = self.cursor.fetchall()
        
        kitap_var = tk.StringVar()
        kitap_combo = ttk.Combobox(top, textvariable=kitap_var, state="readonly", width=35)
        kitap_combo['values'] = [f"{k[0]} - {k[1][:40]}" for k in kitaplar]
        kitap_combo.pack(padx=30, pady=5)
        
        # Öğrenci bilgileri
        tk.Label(top, text="Öğrenci Adı:", bg=self.bg_color).pack(anchor="w", padx=30, pady=(10,0))
        ogrenci_entry = ttk.Entry(top, width=38)
        ogrenci_entry.pack(padx=30, pady=5)
        
        tk.Label(top, text="Okul No:", bg=self.bg_color).pack(anchor="w", padx=30)
        no_entry = ttk.Entry(top, width=38)
        no_entry.pack(padx=30, pady=5)
        
        def kaydet():
            if not kitap_var.get() or not ogrenci_entry.get():
                messagebox.showwarning("Eksik", "Tüm alanları doldurun.", parent=top)
                return
            
            kitap_id = int(kitap_var.get().split(" - ")[0])
            self.cursor.execute("""
                INSERT INTO rezervasyonlar (kitap_id, ogrenci_no, ogrenci_ad, tarih, durum)
                VALUES (?, ?, ?, ?, 'Bekliyor')
            """, (kitap_id, no_entry.get(), ogrenci_entry.get(), datetime.now().strftime("%d.%m.%Y")))
            self.conn.commit()
            
            top.destroy()
            self.durum_yaz("Rezervasyon oluşturuldu.")
            messagebox.showinfo("Başarılı", "Rezervasyon kaydedildi.")
        
        tk.Button(top, text="✅ KAYDET", command=kaydet, font=("Segoe UI", 11, "bold"),
                  bg=self.action_color, fg="white", width=15).pack(pady=20)
    
    # --- HIZLI REZERVASYON ---
    def hizli_rezervasyon(self):
        """Seçili kitap için hızlı rezervasyon"""
        secili = self.tree.selection()
        if not secili:
            messagebox.showwarning("Seçim Yok", "Lütfen bir kitap seçin.")
            return
        
        item = self.tree.item(secili)
        kitap_id = item['values'][0]
        kitap_ad = item['values'][1]
        
        top = tk.Toplevel(self.root)
        top.title("📅 Rezervasyon")
        top.geometry("350x250")
        top.configure(bg=self.bg_color)
        
        tk.Label(top, text=f"'{kitap_ad[:30]}...'", font=("Segoe UI", 11, "bold"),
                 bg=self.bg_color, wraplength=300).pack(pady=15)
        tk.Label(top, text="için rezervasyon", bg=self.bg_color).pack()
        
        tk.Label(top, text="Adınız:", bg=self.bg_color).pack(anchor="w", padx=30, pady=(15,0))
        ad_entry = ttk.Entry(top, width=30)
        ad_entry.pack(padx=30)
        
        tk.Label(top, text="Okul Numaranız:", bg=self.bg_color).pack(anchor="w", padx=30, pady=(10,0))
        no_entry = ttk.Entry(top, width=30)
        no_entry.pack(padx=30)
        
        def kaydet():
            if not ad_entry.get() or not no_entry.get():
                messagebox.showwarning("Eksik", "Tüm alanları doldurun.", parent=top)
                return
            
            self.cursor.execute("""
                INSERT INTO rezervasyonlar (kitap_id, ogrenci_no, ogrenci_ad, tarih, durum)
                VALUES (?, ?, ?, ?, 'Bekliyor')
            """, (kitap_id, no_entry.get(), ad_entry.get(), datetime.now().strftime("%d.%m.%Y")))
            self.conn.commit()
            
            top.destroy()
            self.durum_yaz("Sıraya girdiniz.")
            messagebox.showinfo("Başarılı", "Rezervasyon kaydedildi. Kitap müsait olunca size haber verilecek.")
        
        tk.Button(top, text="✅ SIRAYA GİR", command=kaydet, font=("Segoe UI", 10, "bold"),
                  bg=self.action_color, fg="white").pack(pady=20)
    
    # --- ROZET YÖNETİMİ ---
    def rozet_yonetimi(self):
        """Rozet ve puan sistemi yönetimi"""
        top = tk.Toplevel(self.root)
        top.title("🏅 Rozetler ve Puanlar")
        top.geometry("600x500")
        top.configure(bg="#1a1a2e")
        
        tk.Label(top, text="🏅 ROZET VE PUAN SİSTEMİ", font=("Segoe UI", 16, "bold"),
                 bg="#1a1a2e", fg="#f59e0b").pack(pady=15)
        
        # Rozet tanımları
        rozetler = [
            ("📖 İlk Adım", "İlk kitabını ödünç alan"),
            ("📚 Kitap Kurdu", "5 kitap okuyan"),
            ("🏆 Şampiyon", "10 kitap okuyan"),
            ("⚡ Hızlı Okuyucu", "7 günde kitap bitiren"),
            ("🌟 Süper Star", "20 kitap okuyan"),
            ("👑 Efsane", "50 kitap okuyan"),
        ]
        
        tk.Label(top, text="Kazanılabilir Rozetler:", font=("Segoe UI", 12, "bold"),
                 bg="#1a1a2e", fg="white").pack(pady=10)
        
        rozet_frame = tk.Frame(top, bg="#16213e")
        rozet_frame.pack(fill=tk.X, padx=20, pady=10)
        
        for rozet, aciklama in rozetler:
            row = tk.Frame(rozet_frame, bg="#16213e")
            row.pack(fill=tk.X, pady=3)
            tk.Label(row, text=rozet, font=("Segoe UI", 11), bg="#16213e", fg="#f59e0b", width=15, anchor="w").pack(side=tk.LEFT, padx=10)
            tk.Label(row, text=aciklama, font=("Segoe UI", 10), bg="#16213e", fg="#aaa").pack(side=tk.LEFT)
        
        tk.Label(top, text="\nPuan Sistemi:", font=("Segoe UI", 12, "bold"),
                 bg="#1a1a2e", fg="white").pack(pady=10)
        
        puan_frame = tk.Frame(top, bg="#16213e")
        puan_frame.pack(fill=tk.X, padx=20, pady=10)
        
        puanlar = [
            ("Kitap ödünç alma", "+10 puan"),
            ("7 gün içinde iade", "+5 bonus"),
            ("14 gün içinde iade", "+3 bonus"),
            ("Gecikme", "-2 puan/gün"),
        ]
        
        for islem, puan in puanlar:
            row = tk.Frame(puan_frame, bg="#16213e")
            row.pack(fill=tk.X, pady=3)
            tk.Label(row, text=islem, font=("Segoe UI", 10), bg="#16213e", fg="white", width=20, anchor="w").pack(side=tk.LEFT, padx=10)
            renk = "#10b981" if "+" in puan else "#ef4444"
            tk.Label(row, text=puan, font=("Segoe UI", 10, "bold"), bg="#16213e", fg=renk).pack(side=tk.LEFT)
    
    # --- ÖĞRENCİ GEÇMİŞİ ---
    def ogrenci_gecmisi(self, okul_no):
        """Öğrencinin kitap okuma geçmişi"""
        top = tk.Toplevel(self.root)
        top.title(f"📚 Öğrenci Geçmişi")
        top.geometry("600x500")
        top.configure(bg="#1a1a2e")
        
        # Öğrenci bilgisi
        self.cursor.execute("SELECT ad_soyad, sinif FROM ogrenciler WHERE okul_no=?", (okul_no,))
        ogrenci = self.cursor.fetchone()
        
        if ogrenci:
            tk.Label(top, text=f"👤 {ogrenci[0]} ({ogrenci[1]})", font=("Segoe UI", 14, "bold"),
                     bg="#1a1a2e", fg="#3b82f6").pack(pady=15)
        
        # İstatistikler
        self.cursor.execute("SELECT COUNT(*) FROM odunc_gecmisi WHERE ogrenci_no=?", (okul_no,))
        toplam = self.cursor.fetchone()[0]
        
        stat_frame = tk.Frame(top, bg="#16213e")
        stat_frame.pack(fill=tk.X, padx=20, pady=10)
        
        tk.Label(stat_frame, text=f"📚 Toplam Okunan: {toplam} kitap", font=("Segoe UI", 11),
                 bg="#16213e", fg="white").pack(pady=5)
        
        # Geçmiş listesi
        tk.Label(top, text="Okuma Geçmişi:", font=("Segoe UI", 12, "bold"),
                 bg="#1a1a2e", fg="white").pack(pady=10)
        
        tree_frame = tk.Frame(top, bg="#16213e")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        columns = ("Kitap", "Yazar", "AlinmaTarihi", "IadeTarihi")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=12)
        
        tree.heading("Kitap", text="Kitap")
        tree.heading("Yazar", text="Yazar")
        tree.heading("AlinmaTarihi", text="Alınma")
        tree.heading("IadeTarihi", text="İade")
        
        tree.column("Kitap", width=200)
        tree.column("Yazar", width=150)
        tree.column("AlinmaTarihi", width=90, anchor="center")
        tree.column("IadeTarihi", width=90, anchor="center")
        
        tree.pack(fill=tk.BOTH, expand=True)
        
        self.cursor.execute("""
            SELECT kitap_ad, yazar, alinma_tarihi, iade_tarihi 
            FROM odunc_gecmisi 
            WHERE ogrenci_no=? 
            ORDER BY id DESC
        """, (okul_no,))
        
        for row in self.cursor.fetchall():
            tree.insert("", tk.END, values=row)
    
    # ========================================
    # OYUNLAR (Öğrenci Eğlence)
    # ========================================
    
    def oyun_pong(self):
        """Pong oyunu - Bilgisayara karşı"""
        top = tk.Toplevel(self.root)
        top.title("🏓 Pong")
        top.geometry("600x450")
        top.configure(bg="#000")
        top.resizable(False, False)
        
        canvas = tk.Canvas(top, width=600, height=400, bg="#1a1a2e", highlightthickness=0)
        canvas.pack()
        
        # Oyun değişkenleri
        game = {
            'running': True,
            'player_y': 170,
            'cpu_y': 170,
            'ball_x': 290,
            'ball_y': 190,
            'ball_dx': 4,
            'ball_dy': 3,
            'player_score': 0,
            'cpu_score': 0
        }
        
        # Çizimler
        player = canvas.create_rectangle(20, game['player_y'], 30, game['player_y'] + 60, fill="#3b82f6")
        cpu = canvas.create_rectangle(570, game['cpu_y'], 580, game['cpu_y'] + 60, fill="#ef4444")
        ball = canvas.create_oval(game['ball_x'], game['ball_y'], game['ball_x'] + 20, game['ball_y'] + 20, fill="white")
        score_text = canvas.create_text(300, 30, text="0 - 0", fill="white", font=("Arial", 20, "bold"))
        canvas.create_line(300, 0, 300, 400, fill="#333", dash=(4, 4))
        
        def move_player(event):
            if event.keysym == 'Up' and game['player_y'] > 10:
                game['player_y'] -= 20
            elif event.keysym == 'Down' and game['player_y'] < 330:
                game['player_y'] += 20
            canvas.coords(player, 20, game['player_y'], 30, game['player_y'] + 60)
        
        def game_loop():
            if not game['running']:
                return
            
            # Top hareketi
            game['ball_x'] += game['ball_dx']
            game['ball_y'] += game['ball_dy']
            
            # Üst/alt duvar çarpması
            if game['ball_y'] <= 0 or game['ball_y'] >= 380:
                game['ball_dy'] *= -1
            
            # Raket çarpması - oyuncu
            if 20 <= game['ball_x'] <= 30:
                if game['player_y'] <= game['ball_y'] + 10 <= game['player_y'] + 60:
                    game['ball_dx'] *= -1
                    game['ball_dx'] = min(game['ball_dx'] + 0.5, 10)
            
            # Raket çarpması - bilgisayar
            if 570 <= game['ball_x'] + 20 <= 580:
                if game['cpu_y'] <= game['ball_y'] + 10 <= game['cpu_y'] + 60:
                    game['ball_dx'] *= -1
            
            # Gol
            if game['ball_x'] < 0:
                game['cpu_score'] += 1
                game['ball_x'], game['ball_y'] = 290, 190
                game['ball_dx'] = 4
            elif game['ball_x'] > 580:
                game['player_score'] += 1
                game['ball_x'], game['ball_y'] = 290, 190
                game['ball_dx'] = -4
            
            # Bilgisayar hareketi
            if game['cpu_y'] + 30 < game['ball_y'] and game['cpu_y'] < 330:
                game['cpu_y'] += 3
            elif game['cpu_y'] + 30 > game['ball_y'] and game['cpu_y'] > 10:
                game['cpu_y'] -= 3
            
            # Güncelle
            canvas.coords(ball, game['ball_x'], game['ball_y'], game['ball_x'] + 20, game['ball_y'] + 20)
            canvas.coords(cpu, 570, game['cpu_y'], 580, game['cpu_y'] + 60)
            canvas.itemconfig(score_text, text=f"{game['player_score']} - {game['cpu_score']}")
            
            if game['player_score'] >= 5 or game['cpu_score'] >= 5:
                winner = "Kazandınız! 🎉" if game['player_score'] >= 5 else "Kaybettiniz 😢"
                canvas.create_text(300, 200, text=winner, fill="yellow", font=("Arial", 24, "bold"))
                game['running'] = False
            else:
                top.after(16, game_loop)
        
        def close():
            game['running'] = False
            top.destroy()
        
        top.bind("<Up>", move_player)
        top.bind("<Down>", move_player)
        top.protocol("WM_DELETE_WINDOW", close)
        
        tk.Label(top, text="⬆️⬇️ Ok tuşları ile kontrol", bg="#000", fg="#888", font=("Arial", 10)).pack()
        game_loop()
    
    def oyun_flappy(self):
        """Flappy Bird benzeri oyun"""
        top = tk.Toplevel(self.root)
        top.title("🐦 Flappy Bird")
        top.geometry("400x550")
        top.configure(bg="#000")
        top.resizable(False, False)
        
        canvas = tk.Canvas(top, width=400, height=500, bg="#70c5ce", highlightthickness=0)
        canvas.pack()
        
        game = {
            'running': True,
            'bird_y': 250,
            'velocity': 0,
            'gravity': 0.5,
            'pipes': [],
            'score': 0,
            'frame': 0
        }
        
        # Kuş
        bird = canvas.create_oval(80, game['bird_y'], 110, game['bird_y'] + 30, fill="#f59e0b", outline="#d97706", width=2)
        score_text = canvas.create_text(200, 50, text="0", fill="white", font=("Arial", 32, "bold"))
        
        # Zemin
        canvas.create_rectangle(0, 470, 400, 500, fill="#8b4513")
        
        def jump(event):
            if game['running']:
                game['velocity'] = -8
        
        def add_pipe():
            gap = 150
            top_height = random.randint(80, 280)
            pipe_top = canvas.create_rectangle(400, 0, 450, top_height, fill="#2ecc71", outline="#27ae60", width=2)
            pipe_bottom = canvas.create_rectangle(400, top_height + gap, 450, 470, fill="#2ecc71", outline="#27ae60", width=2)
            game['pipes'].append([pipe_top, pipe_bottom, False])  # False = scored
        
        def game_loop():
            if not game['running']:
                return
            
            game['frame'] += 1
            
            # Yerçekimi
            game['velocity'] += game['gravity']
            game['bird_y'] += game['velocity']
            canvas.coords(bird, 80, game['bird_y'], 110, game['bird_y'] + 30)
            
            # Yeni boru ekle
            if game['frame'] % 90 == 0:
                add_pipe()
            
            # Boruları hareket ettir
            for pipe in game['pipes'][:]:
                canvas.move(pipe[0], -3, 0)
                canvas.move(pipe[1], -3, 0)
                
                coords = canvas.coords(pipe[0])
                if coords and coords[2] < 0:
                    canvas.delete(pipe[0])
                    canvas.delete(pipe[1])
                    game['pipes'].remove(pipe)
                
                # Skor
                if coords and coords[2] < 80 and not pipe[2]:
                    game['score'] += 1
                    pipe[2] = True
                    canvas.itemconfig(score_text, text=str(game['score']))
                
                # Çarpışma
                if coords:
                    top_coords = canvas.coords(pipe[0])
                    bottom_coords = canvas.coords(pipe[1])
                    
                    if top_coords and 80 < top_coords[2] and 110 > top_coords[0]:
                        if game['bird_y'] < top_coords[3] or game['bird_y'] + 30 > bottom_coords[1]:
                            game_over()
                            return
            
            # Zemin/tavan çarpışması
            if game['bird_y'] > 440 or game['bird_y'] < 0:
                game_over()
                return
            
            top.after(16, game_loop)
        
        def game_over():
            game['running'] = False
            canvas.create_rectangle(100, 180, 300, 280, fill="#1a1a2e")
            canvas.create_text(200, 210, text="Game Over!", fill="white", font=("Arial", 18, "bold"))
            canvas.create_text(200, 250, text=f"Skor: {game['score']}", fill="#f59e0b", font=("Arial", 14))
        
        def close():
            game['running'] = False
            top.destroy()
        
        top.bind("<space>", jump)
        top.bind("<Up>", jump)
        top.protocol("WM_DELETE_WINDOW", close)
        
        tk.Label(top, text="SPACE veya ⬆️ ile zıpla", bg="#000", fg="#888", font=("Arial", 10)).pack()
        top.after(500, add_pipe)
        game_loop()
    
    def oyun_yilan(self):
        """Yılan oyunu"""
        top = tk.Toplevel(self.root)
        top.title("🐍 Yılan")
        top.geometry("420x470")
        top.configure(bg="#000")
        top.resizable(False, False)
        
        canvas = tk.Canvas(top, width=400, height=400, bg="#1a1a2e", highlightthickness=0)
        canvas.pack(padx=10, pady=10)
        
        cell_size = 20
        game = {
            'running': True,
            'snake': [(100, 100), (80, 100), (60, 100)],
            'direction': 'Right',
            'food': None,
            'score': 0
        }
        
        score_text = canvas.create_text(200, 15, text="Skor: 0", fill="white", font=("Arial", 14, "bold"))
        
        def spawn_food():
            while True:
                x = random.randint(0, 19) * cell_size
                y = random.randint(1, 19) * cell_size
                if (x, y) not in game['snake']:
                    game['food'] = (x, y)
                    break
        
        spawn_food()
        
        def change_direction(event):
            key = event.keysym
            opposites = {'Up': 'Down', 'Down': 'Up', 'Left': 'Right', 'Right': 'Left'}
            if key in opposites and opposites[key] != game['direction']:
                game['direction'] = key
        
        def game_loop():
            if not game['running']:
                return
            
            # Yeni kafa pozisyonu
            head = game['snake'][0]
            if game['direction'] == 'Up':
                new_head = (head[0], head[1] - cell_size)
            elif game['direction'] == 'Down':
                new_head = (head[0], head[1] + cell_size)
            elif game['direction'] == 'Left':
                new_head = (head[0] - cell_size, head[1])
            else:
                new_head = (head[0] + cell_size, head[1])
            
            # Duvar çarpışması
            if new_head[0] < 0 or new_head[0] >= 400 or new_head[1] < 0 or new_head[1] >= 400:
                game_over()
                return
            
            # Kendine çarpma
            if new_head in game['snake']:
                game_over()
                return
            
            game['snake'].insert(0, new_head)
            
            # Yem yeme
            if new_head == game['food']:
                game['score'] += 10
                canvas.itemconfig(score_text, text=f"Skor: {game['score']}")
                spawn_food()
            else:
                game['snake'].pop()
            
            # Çiz
            canvas.delete("snake", "food")
            
            for i, (x, y) in enumerate(game['snake']):
                color = "#10b981" if i == 0 else "#059669"
                canvas.create_rectangle(x, y, x + cell_size - 2, y + cell_size - 2, fill=color, tags="snake")
            
            if game['food']:
                fx, fy = game['food']
                canvas.create_oval(fx + 2, fy + 2, fx + cell_size - 2, fy + cell_size - 2, fill="#ef4444", tags="food")
            
            top.after(100, game_loop)
        
        def game_over():
            game['running'] = False
            canvas.create_rectangle(100, 150, 300, 250, fill="#1a1a2e", outline="#3b82f6", width=2)
            canvas.create_text(200, 180, text="Game Over!", fill="white", font=("Arial", 18, "bold"))
            canvas.create_text(200, 220, text=f"Skor: {game['score']}", fill="#f59e0b", font=("Arial", 14))
        
        def close():
            game['running'] = False
            top.destroy()
        
        top.bind("<Up>", change_direction)
        top.bind("<Down>", change_direction)
        top.bind("<Left>", change_direction)
        top.bind("<Right>", change_direction)
        top.protocol("WM_DELETE_WINDOW", close)
        
        tk.Label(top, text="Ok tuşları ile yönlendir", bg="#000", fg="#888", font=("Arial", 10)).pack()
        game_loop()
    
    # ========================================
    # ALIŞTIRMALAR (Beyin Egzersizleri)
    # ========================================
    
    def alistirma_matematik(self):
        """Hızlı matematik alıştırması"""
        top = tk.Toplevel(self.root)
        top.title("🔢 Hızlı Matematik")
        top.geometry("400x350")
        top.configure(bg="#1e293b")
        top.resizable(False, False)
        
        game = {'score': 0, 'question': 0, 'total': 10, 'running': True}
        
        tk.Label(top, text="🔢 Hızlı Matematik", font=("Segoe UI", 16, "bold"),
                 bg="#1e293b", fg="white").pack(pady=15)
        
        soru_label = tk.Label(top, text="", font=("Segoe UI", 28, "bold"),
                              bg="#1e293b", fg="#3b82f6")
        soru_label.pack(pady=20)
        
        cevap_entry = ttk.Entry(top, font=("Segoe UI", 18), width=10, justify="center")
        cevap_entry.pack(pady=10)
        
        sonuc_label = tk.Label(top, text=f"Soru: 0/{game['total']} | Doğru: 0",
                               bg="#1e293b", fg="#94a3b8", font=("Segoe UI", 11))
        sonuc_label.pack(pady=10)
        
        dogru_cevap = [0]
        
        def yeni_soru():
            if game['question'] >= game['total']:
                oyun_bitti()
                return
            
            game['question'] += 1
            islem = random.choice(['+', '-', '*'])
            
            if islem == '*':
                a, b = random.randint(2, 12), random.randint(2, 10)
            else:
                a, b = random.randint(10, 99), random.randint(10, 99)
            
            dogru_cevap[0] = eval(f"{a}{islem}{b}")
            soru_label.config(text=f"{a} {islem} {b} = ?")
            sonuc_label.config(text=f"Soru: {game['question']}/{game['total']} | Doğru: {game['score']}")
            cevap_entry.delete(0, tk.END)
            cevap_entry.focus()
        
        def kontrol(event=None):
            if not game['running']:
                return
            try:
                cevap = int(cevap_entry.get())
                if cevap == dogru_cevap[0]:
                    game['score'] += 1
            except:
                pass
            yeni_soru()
        
        def oyun_bitti():
            game['running'] = False
            skor_text = f"🎉 Sonuç: {game['score']}/{game['total']}"
            soru_label.config(text=skor_text, fg="#10b981" if game['score'] >= 7 else "#f59e0b")
            cevap_entry.config(state="disabled")
        
        cevap_entry.bind("<Return>", kontrol)
        yeni_soru()
        
        tk.Label(top, text="Cevabı yazıp Enter'a basın", bg="#1e293b", fg="#64748b",
                 font=("Segoe UI", 9)).pack(pady=10)
    
    def alistirma_kelime(self):
        """İngilizce-Türkçe kelime eşleştirme"""
        top = tk.Toplevel(self.root)
        top.title("🔤 İngilizce Kelime")
        top.geometry("450x400")
        top.configure(bg="#1e293b")
        top.resizable(False, False)
        
        kelimeler = [
            ("book", "kitap"), ("library", "kütüphane"), ("student", "öğrenci"),
            ("teacher", "öğretmen"), ("school", "okul"), ("read", "okumak"),
            ("write", "yazmak"), ("learn", "öğrenmek"), ("friend", "arkadaş"),
            ("knowledge", "bilgi"), ("science", "bilim"), ("history", "tarih"),
            ("mathematics", "matematik"), ("language", "dil"), ("computer", "bilgisayar"),
            ("world", "dünya"), ("country", "ülke"), ("city", "şehir"),
            ("family", "aile"), ("water", "su"), ("food", "yemek"),
            ("time", "zaman"), ("day", "gün"), ("night", "gece"), ("morning", "sabah")
        ]
        
        game = {'score': 0, 'question': 0, 'total': 10}
        
        tk.Label(top, text="🔤 İngilizce Kelime", font=("Segoe UI", 16, "bold"),
                 bg="#1e293b", fg="white").pack(pady=15)
        
        ing_label = tk.Label(top, text="", font=("Segoe UI", 24, "bold"),
                             bg="#1e293b", fg="#3b82f6")
        ing_label.pack(pady=15)
        
        tk.Label(top, text="Türkçe karşılığı nedir?", bg="#1e293b", fg="#94a3b8",
                 font=("Segoe UI", 11)).pack()
        
        btn_frame = tk.Frame(top, bg="#1e293b")
        btn_frame.pack(pady=20)
        
        sonuc_label = tk.Label(top, text=f"Soru: 0/{game['total']} | Doğru: 0",
                               bg="#1e293b", fg="#94a3b8", font=("Segoe UI", 11))
        sonuc_label.pack(pady=10)
        
        secenekler = []
        dogru_idx = [0]
        
        def yeni_soru():
            if game['question'] >= game['total']:
                oyun_bitti()
                return
            
            game['question'] += 1
            random.shuffle(kelimeler)
            dogru = kelimeler[0]
            yanlis = random.sample(kelimeler[1:], 3)
            
            ing_label.config(text=dogru[0].upper())
            
            cevaplar = [dogru[1]] + [k[1] for k in yanlis]
            random.shuffle(cevaplar)
            dogru_idx[0] = cevaplar.index(dogru[1])
            
            for i, btn in enumerate(secenekler):
                btn.config(text=cevaplar[i], bg="#334155", fg="white")
            
            sonuc_label.config(text=f"Soru: {game['question']}/{game['total']} | Doğru: {game['score']}")
        
        def cevapla(idx):
            if idx == dogru_idx[0]:
                game['score'] += 1
                secenekler[idx].config(bg="#10b981")
            else:
                secenekler[idx].config(bg="#ef4444")
                secenekler[dogru_idx[0]].config(bg="#10b981")
            top.after(500, yeni_soru)
        
        def oyun_bitti():
            skor = f"🎉 Sonuç: {game['score']}/{game['total']}"
            ing_label.config(text=skor, fg="#10b981" if game['score'] >= 7 else "#f59e0b")
            for btn in secenekler:
                btn.config(state="disabled")
        
        for i in range(4):
            btn = tk.Button(btn_frame, text="", font=("Segoe UI", 12), width=15,
                           bg="#334155", fg="white", cursor="hand2",
                           command=lambda x=i: cevapla(x))
            btn.grid(row=i//2, column=i%2, padx=5, pady=5)
            secenekler.append(btn)
        
        yeni_soru()
    
    def alistirma_hafiza(self):
        """Sayı dizisi hafıza oyunu"""
        top = tk.Toplevel(self.root)
        top.title("🧩 Hafıza Oyunu")
        top.geometry("400x350")
        top.configure(bg="#1e293b")
        top.resizable(False, False)
        
        game = {'level': 1, 'sequence': [], 'showing': False}
        
        tk.Label(top, text="🧩 Hafıza Oyunu", font=("Segoe UI", 16, "bold"),
                 bg="#1e293b", fg="white").pack(pady=15)
        
        tk.Label(top, text="Sayıları ezberle ve sırayla yaz!", bg="#1e293b", fg="#94a3b8",
                 font=("Segoe UI", 10)).pack()
        
        sayi_label = tk.Label(top, text="Hazır mısın?", font=("Segoe UI", 32, "bold"),
                              bg="#1e293b", fg="#3b82f6")
        sayi_label.pack(pady=25)
        
        level_label = tk.Label(top, text="Seviye: 1", bg="#1e293b", fg="#f59e0b",
                               font=("Segoe UI", 12, "bold"))
        level_label.pack()
        
        cevap_entry = ttk.Entry(top, font=("Segoe UI", 16), width=15, justify="center")
        cevap_entry.pack(pady=15)
        cevap_entry.config(state="disabled")
        
        def yeni_tur():
            game['sequence'] = [random.randint(0, 9) for _ in range(game['level'] + 2)]
            game['showing'] = True
            cevap_entry.config(state="disabled")
            goster_sayilar(0)
        
        def goster_sayilar(idx):
            if idx < len(game['sequence']):
                sayi_label.config(text=str(game['sequence'][idx]))
                top.after(800, lambda: goster_sayilar(idx + 1))
            else:
                sayi_label.config(text="?")
                game['showing'] = False
                cevap_entry.config(state="normal")
                cevap_entry.delete(0, tk.END)
                cevap_entry.focus()
        
        def kontrol(event=None):
            if game['showing']:
                return
            
            cevap = cevap_entry.get().replace(" ", "")
            dogru = "".join(map(str, game['sequence']))
            
            if cevap == dogru:
                game['level'] += 1
                level_label.config(text=f"Seviye: {game['level']}")
                sayi_label.config(text="✓ Doğru!", fg="#10b981")
                top.after(1000, lambda: [sayi_label.config(fg="#3b82f6"), yeni_tur()])
            else:
                sayi_label.config(text=f"✗ Yanlış! Doğru: {dogru}", fg="#ef4444")
                level_label.config(text=f"Final Seviye: {game['level']}")
                cevap_entry.config(state="disabled")
        
        cevap_entry.bind("<Return>", kontrol)
        
        tk.Button(top, text="▶ BAŞLA", font=("Segoe UI", 11, "bold"),
                  bg="#3b82f6", fg="white", cursor="hand2",
                  command=yeni_tur).pack(pady=10)
    
    def alistirma_karistirma(self):
        """Harf karıştırma - kelimeyi bul"""
        top = tk.Toplevel(self.root)
        top.title("📝 Harf Karıştırma")
        top.geometry("400x350")
        top.configure(bg="#1e293b")
        top.resizable(False, False)
        
        kelimeler = ["kütüphane", "öğrenci", "matematik", "bilgisayar", "öğretmen",
                     "kitaplık", "okumak", "yazmak", "düşünmek", "öğrenmek",
                     "arkadaş", "dostluk", "başarı", "çalışmak", "eğitim"]
        
        game = {'score': 0, 'question': 0, 'total': 8, 'kelime': ''}
        
        tk.Label(top, text="📝 Harf Karıştırma", font=("Segoe UI", 16, "bold"),
                 bg="#1e293b", fg="white").pack(pady=15)
        
        tk.Label(top, text="Harfleri düzenleyerek kelimeyi bul!", bg="#1e293b", fg="#94a3b8",
                 font=("Segoe UI", 10)).pack()
        
        karisik_label = tk.Label(top, text="", font=("Segoe UI", 28, "bold"),
                                 bg="#1e293b", fg="#f59e0b")
        karisik_label.pack(pady=20)
        
        cevap_entry = ttk.Entry(top, font=("Segoe UI", 16), width=15, justify="center")
        cevap_entry.pack(pady=10)
        
        sonuc_label = tk.Label(top, text=f"Soru: 0/{game['total']} | Doğru: 0",
                               bg="#1e293b", fg="#94a3b8", font=("Segoe UI", 11))
        sonuc_label.pack(pady=10)
        
        def yeni_soru():
            if game['question'] >= game['total']:
                oyun_bitti()
                return
            
            game['question'] += 1
            game['kelime'] = random.choice(kelimeler)
            harfler = list(game['kelime'])
            random.shuffle(harfler)
            karisik_label.config(text=" ".join(harfler).upper())
            sonuc_label.config(text=f"Soru: {game['question']}/{game['total']} | Doğru: {game['score']}")
            cevap_entry.delete(0, tk.END)
            cevap_entry.focus()
        
        def kontrol(event=None):
            cevap = cevap_entry.get().lower().strip()
            if cevap == game['kelime']:
                game['score'] += 1
            yeni_soru()
        
        def oyun_bitti():
            skor = f"🎉 Sonuç: {game['score']}/{game['total']}"
            karisik_label.config(text=skor, fg="#10b981" if game['score'] >= 5 else "#f59e0b")
            cevap_entry.config(state="disabled")
        
        cevap_entry.bind("<Return>", kontrol)
        yeni_soru()
    
    def alistirma_tepki(self):
        """Hızlı tepki testi"""
        top = tk.Toplevel(self.root)
        top.title("⚡ Hızlı Tepki")
        top.geometry("400x380")
        top.configure(bg="#1e293b")
        top.resizable(False, False)
        
        game = {'waiting': False, 'start_time': 0, 'results': [], 'round': 0}
        
        tk.Label(top, text="⚡ Hızlı Tepki Testi", font=("Segoe UI", 16, "bold"),
                 bg="#1e293b", fg="white").pack(pady=15)
        
        tk.Label(top, text="Yeşil olunca hemen tıkla!", bg="#1e293b", fg="#94a3b8",
                 font=("Segoe UI", 10)).pack()
        
        btn = tk.Button(top, text="BAŞLAMAK İÇİN TIKLA", font=("Segoe UI", 14, "bold"),
                        width=25, height=4, bg="#334155", fg="white", cursor="hand2")
        btn.pack(pady=25)
        
        sonuc_label = tk.Label(top, text="", bg="#1e293b", fg="#94a3b8",
                               font=("Segoe UI", 12))
        sonuc_label.pack(pady=10)
        
        ort_label = tk.Label(top, text="", bg="#1e293b", fg="#f59e0b",
                             font=("Segoe UI", 11, "bold"))
        ort_label.pack()
        
        def bekle():
            game['round'] += 1
            btn.config(text="BEKLE...", bg="#ef4444", fg="white")
            game['waiting'] = False
            
            bekleme = random.randint(1500, 4000)
            top.after(bekleme, yesil_yap)
        
        def yesil_yap():
            if not game['waiting']:
                game['waiting'] = True
                game['start_time'] = time.time()
                btn.config(text="TIKLA!", bg="#10b981")
        
        def tikla():
            if game['waiting']:
                tepki = int((time.time() - game['start_time']) * 1000)
                game['results'].append(tepki)
                sonuc_label.config(text=f"⏱️ {tepki} ms", fg="#10b981" if tepki < 300 else "#f59e0b")
                
                if game['round'] >= 5:
                    ort = sum(game['results']) // len(game['results'])
                    ort_label.config(text=f"Ortalama: {ort} ms")
                    btn.config(text="🎉 BİTTİ!", bg="#3b82f6", state="disabled")
                else:
                    btn.config(text="DEVAM", bg="#334155")
                    game['waiting'] = False
            elif game['round'] == 0 or btn['text'] == "DEVAM":
                bekle()
            else:
                sonuc_label.config(text="⚠️ Çok erken!", fg="#ef4444")
                btn.config(text="TEKRAR", bg="#334155")
                game['waiting'] = False
        
        btn.config(command=tikla)
    
    # --- OTOMATİK YEDEKLEME ---
    def otomatik_yedekleme_baslat(self):
        """Arka planda otomatik yedekleme"""
        def yedekle_thread():
            while True:
                time.sleep(3600 * 24)  # Günlük yedekleme
                try:
                    base_dir = os.path.dirname(os.path.abspath(__file__))
                    yedek_dir = os.path.join(base_dir, "yedekler")
                    if not os.path.exists(yedek_dir):
                        os.makedirs(yedek_dir)
                    
                    src = os.path.join(base_dir, self.db_adi)
                    dst = os.path.join(yedek_dir, f"otomatik_{datetime.now().strftime('%Y%m%d_%H%M')}.db")
                    shutil.copy2(src, dst)
                    
                    # Eski yedekleri temizle (son 5 tane tut)
                    yedekler = sorted([f for f in os.listdir(yedek_dir) if f.startswith("otomatik_")])
                    while len(yedekler) > 5:
                        os.remove(os.path.join(yedek_dir, yedekler.pop(0)))
                except:
                    pass
        
        thread = threading.Thread(target=yedekle_thread, daemon=True)
        thread.start()
    
    # --- ÇIKIŞ YAP ---
    def cikis_yap(self):
        """Oturumu kapat ve giriş ekranına dön"""
        if messagebox.askyesno("Çıkış", "Oturumu kapatmak istiyor musunuz?"):
            self.root.destroy()
            if CTK_DESTEGI:
                yeni_root = ctk.CTk()
            else:
                yeni_root = tk.Tk()
            GirisEkrani(yeni_root, lambda tip: ana_uygulama_baslat(yeni_root, tip))
            yeni_root.mainloop()


def ana_uygulama_baslat(giris_root, kullanici_tipi):
    """Giriş ekranını kapat ve ana uygulamayı başlat"""
    giris_root.destroy()
    root = tk.Tk()
    app = KutuphaneUygulamasi(root, kullanici_tipi)
    root.mainloop()


if __name__ == "__main__":
    if CTK_DESTEGI:
        root = ctk.CTk()
    else:
        root = tk.Tk()
    GirisEkrani(root, lambda tip: ana_uygulama_baslat(root, tip))
    root.mainloop()